#!/usr/bin/env python3

from datetime import datetime
from easy_functions import process_kwargs
from easy_functions import process_method
from intersight.api import asset_api
from intersight.api import compute_api
from intersight.api import fcpool_api
from intersight.api import server_api
from intersight.api import vnic_api
from operator import itemgetter
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from pathlib import Path
import credentials
import jinja2
import json
import pkg_resources
import pytz
import openpyxl
import os
import re
import validating
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

home = Path.home()
template_path = pkg_resources.resource_filename('class_vmware', 'Templates/')

# Exception Classes
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

class servers(object):
    def __init__(self, type):
        # self.templateLoader = jinja2.FileSystemLoader(
        #     searchpath=(template_path + f'{ws}/'))
        # self.templateEnv = jinja2.Environment(loader=self.templateLoader)
        # self.name_prefix = name_prefix
        # self.org = org
        self.type = type


    def server_inventory(self, **kwargs):
        args = kwargs['args']
        pyDict = {}
        # Obtain Server Profile Data
        api_client = credentials.config_credentials(home, args)
        api_handle = server_api.ServerApi(api_client)
        apiQuery = api_handle.get_server_profile_list()
        if apiQuery.results:
            apiResults = apiQuery.results
            for i in apiResults:
                profileMoid = i.moid
                profileName = i.name
                if i.target_platform == 'FIAttached':
                    api_handle = fcpool_api.FcpoolApi(api_client)
                    query_filter = f"PoolPurpose eq 'WWNN' and AssignedToEntity.Moid eq '{profileMoid}'"
                    kwargs = dict(filter=query_filter)
                    fcPool = api_handle.get_fcpool_lease_list(**kwargs)
                    wwnn_address = fcPool.results[0].wwn_id
                if i.associated_server:
                    # Obtain Physical UCS Server Information
                    serverMoid = i.associated_server.moid
                    serverType = i.associated_server.object_type
                    api_handle = compute_api.ComputeApi(api_client)
                    if serverType == 'compute.Blade':
                        apiQuery = api_handle.get_compute_blade_by_moid(serverMoid)
                        serverDn = 'chassis-' + str(apiQuery.chassis_id) + "/blade-" + str(apiQuery.slot_id)
                    else:
                        apiQuery = api_handle.get_compute_rack_unit_by_moid(serverMoid)
                        serverDn = 'rackunit-' + str(apiQuery.server_id)
                    serverSerial = apiQuery.serial
                    api_handle = asset_api.AssetApi(api_client)
                    serverReg = api_handle.get_asset_device_registration_by_moid(apiQuery.registered_device.moid)
                    if i.target_platform == 'FIAttached':
                        domainParent = api_handle.get_asset_device_registration_by_moid(serverReg.parent_connection.moid)
                    # Obtain Server vNICs (Ethernet/Fibre-Channel)
                    api_handle = vnic_api.VnicApi(api_client)
                    query_filter = f"Profile.Moid eq '{profileMoid}'"
                    kwargs = dict(filter=query_filter)
                    eth_apiQuery = api_handle.get_vnic_eth_if_list(**kwargs)
                    fc_apiQuery = api_handle.get_vnic_fc_if_list(**kwargs)
                    vnics = {}
                    for item in eth_apiQuery.results:
                        vnic_name = item['name']
                        mac_address = item['mac_address']
                        ngpMoid = item['fabric_eth_network_group_policy'][0].moid
                        qosMoid = item['eth_qos_policy']['moid']
                        api_handle = vnic_api.VnicApi(api_client)
                        qosPolicy = api_handle.get_vnic_eth_qos_policy_by_moid(qosMoid)
                        mTu = qosPolicy.mtu
                        vnic = {
                            vnic_name: {
                                'mac_address':mac_address,
                                'mtu':mTu,
                            }
                        }
                        vnics.update(vnic)

                    vhbas = {'A':{},'B':{}}
                    for item in fc_apiQuery.results:
                        vhba_name = item['name']
                        wwpn_address = item['wwpn']
                        switch_id = item['placement']['switch_id']
                        fcnpMoid = item['fc_network_policy'].moid
                        vhbas[switch_id] = {
                            'vhba':vhba_name,
                            'wwpn_address':wwpn_address
                        }
                    if i.target_platform == 'FIAttached':
                        hostResults = {
                            profileName:{
                                'domain_name':domainParent.device_hostname[0],
                                'moid':profileMoid,
                                'serial':serverSerial,
                                'server_dn':serverDn,
                                'vhbas':vhbas,
                                'vnics':vnics,
                                'wwnn':wwnn_address
                            }
                        }
                    else:
                        hostResults = {
                            profileName:{
                                'domain_name':'N/A',
                                'moid':profileMoid,
                                'serial':serverSerial,
                                'server_dn':serverDn,
                                'vhbas':vhbas,
                                'vnics':vnics,
                                'wwnn':'N/A'
                            }
                        }
                    if i.target_platform == 'FIAttached':
                        pyDict.update(hostResults)

        # print(json.dumps(pyDict, indent=4))
        # Build Named Style Sheets for Workbook
        bd1 = Side(style="thick", color="0070C0")
        bd2 = Side(style="medium", color="0070C0")
        heading_1 = NamedStyle(name="heading_1")
        heading_1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
        heading_1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
        heading_1.fill = PatternFill("solid", fgColor="305496")
        heading_1.font = Font(bold=True, size=15, color="FFFFFF")
        heading_2 = NamedStyle(name="heading_2")
        heading_2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
        heading_2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
        heading_2.font = Font(bold=True, size=15, color="44546A")
        even = NamedStyle(name="even")
        even.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
        even.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
        even.font = Font(bold=False, size=12, color="44546A")
        odd = NamedStyle(name="odd")
        odd.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
        odd.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
        odd.fill = PatternFill("solid", fgColor="D9E1F2")
        odd.font = Font(bold=False, size=12, color="44546A")

        Est = pytz.timezone('US/Eastern')
        datetime_est = datetime.now(Est)
        Est1 = datetime_est.strftime('%Y-%m-%d_%H-%M')
        Est2 = datetime_est.strftime('%Y-%m-%d %H:%M:%S %Z %z')
        workbook = f'UCS_WWPN_Collector-{Est1}.xlsx'
        wb = openpyxl.Workbook()
        wb.add_named_style(heading_1)
        wb.add_named_style(heading_2)
        wb.add_named_style(even)
        wb.add_named_style(odd)
        ws = wb.active
        ws.title = 'WWPN List'
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws_header = f'Collected UCS Data on {Est2}'
        data = [ws_header]
        ws.append(data)
        ws.merge_cells('A1:D1')
        for cell in ws['1:1']:
            cell.style = 'heading_1'
        data = [
            'Server Profile','WWNN','WWPN for Fabric A','WWPN for Fabric B'
        ]
        ws.append(data)
        for cell in ws['2:2']:
            cell.style = 'heading_2'
        ws_row_count = 3
        for k, v in pyDict.items():
            if v['wwnn']: wwnn = v['wwnn']
            else: wwnn = 'Not Configured'
            if v['vhbas']['A'].get('wwpn_address'):
                vhba1_wwpn = v['vhbas']['A']['wwpn_address']
            else:
                vhba1_wwpn = 'Not Configured'
            if v['vhbas']['B'].get('wwpn_address'):
                vhba2_wwpn = v['vhbas']['B']['wwpn_address']
            else:
                vhba2_wwpn = 'Not Configured'
            data = [
                k,
                wwnn,
                vhba1_wwpn,
                vhba2_wwpn
            ]
            ws.append(data)
            for cell in ws[ws_row_count:ws_row_count]:
                if ws_row_count % 2 == 0:
                    cell.style = 'odd'
                else:
                    cell.style = 'even'
            ws_row_count += 1
        wb.save(filename=workbook)


