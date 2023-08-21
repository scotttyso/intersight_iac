#=============================================================================
# Print Color Functions
#=============================================================================
def prCyan(skk):        print("\033[96m {}\033[00m" .format(skk))
def prGreen(skk):       print("\033[92m {}\033[00m" .format(skk))
def prLightPurple(skk): print("\033[94m {}\033[00m" .format(skk))
def prLightGray(skk):   print("\033[94m {}\033[00m" .format(skk))
def prPurple(skk):      print("\033[95m {}\033[00m" .format(skk))
def prRed(skk):         print("\033[91m {}\033[00m" .format(skk))
def prYellow(skk):      print("\033[93m {}\033[00m" .format(skk))

#=============================================================================
# Source Modules
#=============================================================================
try:
    from classes import ezfunctions as ezfunctions
    from classes import isight
    from copy import deepcopy
    from datetime import datetime
    from dotmap import DotMap
    from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
    import numpy, pytz, openpyxl, re, sys, urllib3, yaml
except ImportError as e:
    prRed(f'!!! ERROR !!!\n{e.__class__.__name__}')
    prRed(f" Module {e.name} is required to run this script")
    prRed(f" Install the module using the following: `pip install {e.name}`")
    sys.exit(1)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

class intersight_api(object):
    def __init__(self, type):
        self.type = type

    #======================================================
    # Function - Add Policies
    #======================================================
    def add_policies(self, kwargs):
        #======================================================
        # Check for YAML File Definition
        #======================================================
        prCyan(f'\n{"-"*91}\n  Beginning Policy Append to Profiles...\n{"-"*91}\n')
        if kwargs.args.yaml_file != None:
            yaml_arg = True
            ydata = DotMap(yaml.safe_load(open(kwargs.args.yaml_file, 'r')))
        else: yaml_arg = False; ydata = DotMap()
        #======================================================
        # Build Platform Data
        #======================================================
        target_platforms = ['chassis', 'domain', 'FIAttached', 'Standalone']
        target = DotMap()
        for e in target_platforms:
            target[e].policies = []
        for k, v in kwargs.ez_data.items():
            if v.get('target_platforms') and v.get('intersight_type') == 'policy':
                for e in v.target_platforms:
                    target[e].policies.append(k)
        #======================================================
        # Determine if the Profiles are FIAttached or Standalone
        #======================================================
        target_platforms = ['chassis', 'domain', 'server', 'server_template']
        kwargs.jdata = DotMap()
        if yaml_arg == False:
            kwargs.jdata = DotMap(
                default = 'server', description  = 'Select the Intersight Profile Type.',
                enum = target_platforms,  multi_select = True, title = 'Profile Type(s)')
            profile_types = ezfunctions.variablePrompt(kwargs)
        else: profile_types = [e.type for e in ydata.profile_types]
        kwargs.sub_type = DotMap()
        if yaml_arg == False:
            for e in profile_types:
                if re.search('^server(_template)?$', e):
                    kwargs.jdata = DotMap(
                        default = 'FIAttached', description  = f'Select the {e} Profile Sub-Type.',
                        enum = ['FIAttached', 'Standalone'], title = 'Profile Sub-Type')
                    kwargs.sub_type[e] = ezfunctions.variablePrompt(kwargs)
        else:
            for e in ydata.profile_types:
                if re.search('server|server_template'):
                    kwargs.sub_type[e.type] = e.sub_type
        #======================================================
        # Prompt User for Policies to Attach
        #======================================================
        for e in profile_types:
            if yaml_arg == False:
                if re.search('server(_template)?', e): policies = target[kwargs.sub_type[e]].policies
                else: policies = target[e].policies
                kwargs.jdata = DotMap(
                    default      = policies[0],
                    description  = f'Select the `{", ".join(profile_types)}` Policy Type(s) (can be multiple).',
                    enum         = policies,
                    multi_select = True,
                    title        = 'Policy Type(s)')
                target[e].update_policies = ezfunctions.variablePrompt(kwargs)
            else: target[e].update_policies = ydata.policy_types
        #======================================================
        # Prompt for Organizations and Loop Through Target Platforms
        #======================================================
        if yaml_arg == False:  kwargs = select_organizations(kwargs)
        else: kwargs.organizations = ydata.organizations
        for target_platform in profile_types:
            for org in kwargs.organizations:
                prCyan(f'\n{"-"*91}\n  Starting `{target_platform}` Loop in Organization `{org}`.\n{"-"*91}\n')
                update_profile(org, target, target_platform, yaml_arg, ydata, kwargs)
                prCyan(f'\n{"-"*91}\n  Finished `{target_platform}` Loop in Organization `{org}`.\n{"-"*91}\n')
            prPurple(f'\n{"-"*91}\n  Finished Updating `{target_platform}` Profiles...\n{"-"*91}\n')


    #======================================================
    # Function - Add Vlans
    #======================================================
    def add_vlans(self, kwargs):
        #======================================================
        # Function - Add Vlans
        #======================================================
        # Validate YAML configuration file is defined.
        if kwargs.args.yaml_file != None:
            ydata = DotMap(yaml.safe_load(open(kwargs.args.yaml_file, 'r'))).add_vlans
        else:
            prRed(f'\n{"-"*91}\n\n  Missing Required YAML File Argument `-y`.  Exiting Process.\n\n{"-"*91}\n')
            sys.exit(1)
        #======================================================
        # Get VLAN List and Organizations from YAML configuration file.
        #======================================================
        tags  = [{'key': 'Module','value': 'day2tools'}]
        vlans = ydata.patch.vlan_policy.vlans
        kwargs.organizations = ydata.organizations
        for org in kwargs.organizations:
            kwargs.org = org
            prCyan(f'\n{"-"*91}\n\n  Starting Loop on Organization {org}.\n\n{"-"*91}\n')
            #======================================================
            # Query the API for the VLAN Policies
            #======================================================
            vpolicy  = ydata.patch.vlan_policy.name
            kwargs = isight.api_get(False, [vpolicy], 'vlan', kwargs)
            kwargs.vlans = kwargs.pmoids #880-0551
            #======================================================
            # Query the API for the VLANs Attached to the VLAN Policy
            #======================================================
            prCyan(f'\n{"-"*91}\n\n  Checking VLAN Policy `{vpolicy}` for VLANs.\n\n{"-"*91}\n')
            kwargs.pmoid = kwargs.vlans[vpolicy].moid
            kwargs = isight.api_get(True, [e.vlan_id for e in vlans], 'vlan.vlans', kwargs)
            mcast_moid   = kwargs.results[0].MulticastPolicy.Moid
            policy_vlans = [e.VlanId for e in kwargs.results]
            add_vlans = []
            for e in vlans:
                if not e.vlan_id in policy_vlans: add_vlans.append(e)
                else:
                    prCyan(f'\n{"-"*91}\n')
                    prCyan(f'  VLAN `{e.vlan_id}` is already in VLAN Policy `{vpolicy}` in Organization `{org}`.\n\n{"-"*91}\n')
            for e in add_vlans:
                prCyan(f'\n{"-"*91}\n')
                prCyan(f'  VLAN `{e.vlan_id}` is not in VLAN Policy: `{vpolicy}` in Organization: `{org}`.  Adding VLAN...')
                prCyan(f'\n{"-"*91}\n')
                kwargs.apiBody = {
                    'EthNetworkPolicy':{'Moid':kwargs.vlans[vpolicy].moid,'ObjectType':'fabric.EthNetworkPolicy'},
                    'MulticastPolicy':{'Moid':mcast_moid,'ObjectType':'fabric.MulticastPolicy'},
                    'Name':e.name, 'ObjectType': 'fabric.Vlan', 'VlanId':e.vlan_id}
                kwargs.method = 'post'
                kwargs.uri    = kwargs.ez_data['vlan.vlans'].intersight_uri
                kwargs        = isight.api('vlan').calls(kwargs)
            if len(ydata.patch.ethernet_network_group) > 0:
                #======================================================
                # Query the API for the Ethernet Network Group Policies
                #======================================================
                kwargs = isight.api_get(False, ydata.patch.ethernet_network_group, 'ethernet_network_group', kwargs)
                eng_results  = kwargs.results
                vlan_ids = [v.vlan_id for v in add_vlans]
                for e in ydata.patch.ethernet_network_group:
                    indx = [x for x, i in enumerate(eng_results) if e in i.Name][0]
                    allowed_vlans = ezfunctions.vlan_list_full(eng_results[indx].VlanSettings.AllowedVlans)
                    allowed_vlans = ezfunctions.vlan_list_format(list(numpy.unique(numpy.array(allowed_vlans + vlan_ids))))
                    kwargs.apiBody= {'Name':e,'VlanSettings':{'AllowedVlans':allowed_vlans}}
                    kwargs.method = 'patch'
                    kwargs.pmoid  = eng_results[indx].Moid
                    kwargs        = isight.api(kwargs.qtype).calls(kwargs)
            #======================================================
            # Loop through Policy Creation
            #======================================================
            if len(ydata.post) > 0:
                idata = DotMap(ethernet_adapter  = [], ethernet_network_control = [], ethernet_network_group = [],
                               ethernet_qos = [], lan_connectivity = [], mac = [])
                for e in add_vlans:
                    idata.lan_connectivity.append(DotMap(
                        name = ydata.post.lan_connectivity.name.replace('{{vlan_id}}', str(e.vlan_id)),
                        vlan_id = e.vlan_id))
                    for i in ydata.post.lan_connectivity.vnics:
                        idata.ethernet_adapter.append(i.ethernet_adapter)
                        idata.ethernet_network_control.append(i.ethernet_network_control)
                        idata.ethernet_network_group.append(DotMap(
                            name = i.ethernet_network_group.replace('{{vlan_id}}', str(e.vlan_id)), vlan_id = e.vlan_id))
                        idata.ethernet_qos.append(i.ethernet_qos)
                        idata.mac.append(i.mac_pool)
                for k, v in idata.items():idata[k] = sorted(list(numpy.unique(numpy.array(v))))
                #======================================================
                # Query the API for the Ethernet Network Group Policies
                #======================================================
                kwargs = isight.api_get(True, [e.name for e in idata.ethernet_network_group], 'ethernet_network_group', kwargs)
                eth_eng= kwargs.pmoids
                eng_keys = list(eth_eng.keys())
                for e in idata.ethernet_network_group:
                    if e.name in eng_keys:
                        prCyan(f'\n{"-"*91}')
                        prCyan(f'Ethernet Network Group `{e.name}` Exists.  Moid is: {eth_eng[e.name].moid}')
                        prCyan(f'\n{"-"*91}')
                    else:
                        prCyan(f'\n{"-"*91}\n  Ethernet Network Group `{e.name}` does not exist.  Creating...\n{"-"*91}\n')
                        kwargs.apiBody = {
                            'Description': f'{e.name} Ethernet Network Group', 'Name': e.name,
                            'ObjectType':'fabric.EthNetworkGroupPolicy', 'Tags':tags,
                            'Organization': {'Moid':kwargs.org_moids[org].moid,'ObjectType':'organization.Organization'},
                            'VlanSettings':{'AllowedVlans':f'{e.vlan_id}','NativeVlan':e.vlan_id,'ObjectType':'fabric.VlanSettings'}}
                        kwargs.method= 'post'
                        kwargs       = isight.api(kwargs.qtype).calls(kwargs)
                        eth_eng[e.name].moid = kwargs.pmoid
                #======================================================
                # Query the API for Policies
                #======================================================
                pdata = DotMap()
                ilist = ['ethernet_adapter', 'ethernet_network_control', 'ethernet_qos', 'mac']
                for e in ilist:
                    kwargs = isight.api_get(True, idata[e], e, kwargs)
                    pdata[e] = kwargs.pmoids
                #======================================================
                # Query the API for the LAN Connectivity Policies
                #======================================================
                kwargs.names = [e.name for e in idata.lan_connectivity]
                kwargs = isight.api_get(True, [e.name for e in idata.lan_connectivity], 'lan_connectivity', kwargs)
                lan_policies = kwargs.pmoids
            for e in idata.lan_connectivity:
                #======================================================
                # Configure LAN Connectivity Policies
                #======================================================
                if e.name in list(lan_policies.keys()):
                    prCyan(f'\n{"-"*91}\n  LAN Policy `{e.name}` exists.  Moid is: {lan_policies[e.name].moid}\n{"-"*91}\n')
                    lan_moid = lan_policies[e.name].moid
                else:
                    prCyan(f'\n{"-"*91}\n\n  LAN Policy `{e.name}` does not exist.  Creating...\n\n{"-"*91}\n')
                    kwargs.apiBody = {
                        'Name': str(e.name),
                        'ObjectType': 'vnic.LanConnectivityPolicy',
                        'Organization': {'Moid': kwargs.org_moids[org].moid, 'ObjectType': 'organization.Organization'},
                        'Tags': tags, 'TargetPlatform': ydata.post.lan_connectivity.target_platform
                        }
                    kwargs.method= 'post'
                    kwargs.uri   = kwargs.ez_data.lan_connectivity.intersight_uri
                    kwargs       = isight.api(kwargs.qtype).calls(kwargs)
                    lan_moid     = kwargs.pmoid
                #======================================================
                # Configure vNIC Policies
                #======================================================
                kwargs.pmoid = lan_moid
                kwargs = isight.api_get(True, [i.name for i in ydata.post.lan_connectivity.vnics], 'lan_connectivity.vnics', kwargs)
                vnic_results = kwargs.pmoids
                for i in ydata.post.lan_connectivity.vnics:
                    if i.name in list(vnic_results.keys()):
                        prCyan(f'\n{"-"*91}\n')
                        prCyan(f'  LAN Connectivity `{e.name}` vNIC `{i.name}` exists.  Moid is: {vnic_results[i.name].moid}')
                        prCyan(f'\n{"-"*91}\n')
                    else:
                        prCyan(f'\n{"-"*91}\n')
                        prCyan(f'  vNIC `{i.name}` was not attached to LAN Policy `{e.name}`.  Creating...')
                        prCyan(f'\n{"-"*91}\n')
                        if len(ydata.post.lan_connectivity.vnics) > 1: failover = False
                        else: failover = True
                        eng = i.ethernet_network_group.replace('{{vlan_id}}', str(e.vlan_id))
                        kwargs.apiBody = {
                            'Cdn': {'ObjectType': 'vnic.Cdn', 'Source': 'vnic', 'Value': i.name },
                            'EthAdapterPolicy': {'Moid': pdata['ethernet_adapter'][i.ethernet_adapter].moid,
                                                'ObjectType': 'vnic.EthAdapterPolicy'},
                            'EthQosPolicy': {'Moid': pdata['ethernet_qos'][i.ethernet_qos].moid,
                                             'ObjectType': 'vnic.EthQosPolicy'},
                            'FabricEthNetworkControlPolicy': {
                                'Moid': pdata['ethernet_network_control'][i.ethernet_network_control].moid,
                                'ObjectType': 'fabric.EthNetworkControlPolicy'},
                            'FabricEthNetworkGroupPolicy': [{'Moid': eth_eng[eng].moid,
                                                            'ObjectType': 'fabric.EthNetworkGroupPolicy'}],
                            'FailoverEnabled': failover,
                            'LanConnectivityPolicy': {'Moid': lan_policies[e.name].moid,
                                                        'ObjectType': 'vnic.LanConnectivityPolicy'},
                            'MacAddressType': 'POOL',
                            'MacPool': {'Moid': pdata['mac'][i.mac_pool].moid, 'ObjectType': 'macpool.Pool'},
                            'Name': 'NIC-A',
                            'ObjectType': 'vnic.EthIf',
                            'Order': i.placement.order,
                            'Placement': {'Id': i.placement.slot, 'ObjectType': 'vnic.PlacementSettings', 'PciLink': 0,
                                        'SwitchId': i.placement.switch_id, 'Uplink': 0},
                        }
                        kwargs.method = 'post'
                        kwargs.uri    = kwargs.ez_data['lan_connectivity.vnics'].intersight_uri
                        kwargs        = isight.api('vnics').calls(kwargs)
        prCyan(f'\n{"-"*91}\n\n  Finished Loop on Organization `{org}`.\n\n{"-"*91}\n')
        

    #======================================================
    # Function - HCL Inventory
    #======================================================
    def hcl_inventory(self, kwargs):
        py_dict   = DotMap()
        # Obtain Server Profile Data
        for e in kwargs.yaml_data:
            if 'Cisco' in e.Hostname.Manufacturer:
                py_dict[e.Serial] = DotMap(
                    build      = e.Hostname.Build,
                    cluster    = e.Cluster,
                    domain     = 'Standalone',
                    hostname   = e.Hostname.Name,
                    model      = e.Hostname.Model,
                    server_dn  = 'unknown',
                    srv_profile= 'undefined',
                    toolDate   = e.InstallDate,
                    toolName   = e.Name,
                    toolVer    = e.Version,
                    vcenter    = e.vCenter,
                    version    = e.Hostname.Version,
                )

        kwargs.names  = list(py_dict.keys())
        kwargs.method = 'get'
        kwargs.qtype  = 'serial'
        kwargs.uri    = 'compute/PhysicalSummaries'
        kwargs        = isight.api(kwargs.qtype).calls(kwargs)
        registered_devices  = []
        for e in kwargs.results:
            py_dict[e.Serial].server_dn = (e.Dn).replace('sys/', '')
            py_dict[e.Serial].moid = e.Moid
            if e.get('ServiceProfile'): py_dict[e.Serial].srv_profile = e.ServiceProfile
            if not e.ManagementMode == 'IntersightStandalone':
                py_dict[e.Serial].registered_moid = e.RegisteredDevice.Moid
                registered_devices.append(e.RegisteredDevice.Moid)
        if len(registered_devices) > 0:
            domain_map = DotMap()
            parents    = []
            kwargs.method = 'get'
            kwargs.names  = list(set(registered_devices))
            kwargs.qtype = 'registered_device'
            kwargs.uri   = 'asset/DeviceRegistrations'
            for e in kwargs.results:
                if e.get('ParentConnection'):
                    domain_map[e.Moid].hostname = None
                    domain_map[e.Moid].parent = e.ParentConnection.Moid
                    parents.append(e.ParentConnection.Moid)
                else:
                    domain_map[e.Moid].hostname = e.DeviceHostname
                    domain_map[e.Moid].parent = None
            if len(parents) > 0:
                kwargs.names = list(set(parents))
                kwargs.qtype = 'registered_device'
                kwargs.uri   = 'asset/DeviceRegistrations'
                kwargs       = isight.api(kwargs.qtype)
                for e in kwargs.results:
                    for k, v in domain_map.items():
                        if v.get('parent'):
                            if v.parent == i.Moid: domain_map[k].hostname = e.DeviceHostname
            for k, v in py_dict.items():
                if v.get('registered_moid'): py_dict[k].domain = domain_map[v.registered_moid].hostname

        kwargs.names = [v.moid for k, v in py_dict.items()]
        kwargs.qtype = 'hcl_status'
        kwargs.uri   = 'cond/HclStatuses'
        kwargs       = isight.api(kwargs.qtype).calls(kwargs)
        for e in kwargs.results:
            for k, v in py_dict:
                if v.moid == e.ManagedObject.Moid:
                        py_dict[k].firmware = e.HclFirmwareVersion
                        py_dict[k].status   = e.Status

        if len(py_dict) > 0:
            kwargs = get_local_time(kwargs)

            # Build Named Style Sheets for Workbook
            kwargs = workbook_styles(kwargs)
            workbook = f'HCL-Inventory-{kwargs.time_short}.xlsx'
            wb = kwargs.wb
            ws = wb.active
            ws.title = 'Inventory List'

            # Read Server Inventory to Create Column Headers
            column_headers = [
                'Domain','Model','Serial','Server','Profile','Firmware','vCenter','Cluster','Hostname',
                'ESX Version','ESX Build','HCL Component Status', 'UCS Tools Install Date', 'UCS Tools Version'
            ]
            for i in range(len(column_headers)): ws.column_dimensions[chr(ord('@')+i+1)].width = 30
            cLength = len(column_headers)
            ws_header = f'Collected UCS Data on {kwargs.time_long}'
            data = [ws_header]
            ws.append(data)
            ws.merge_cells(f'A1:{chr(ord("@")+cLength)}1')
            for cell in ws['1:1']: cell.style = 'heading_1'
            ws.append(column_headers)
            for cell in ws['2:2']: cell.style = 'heading_2'
            ws_row_count = 3
            
            # Populate the Columns with Server Inventory
            for key, value in py_dict.items():
                # Add the Columns to the Spreadsheet
                for k, v in value.items(): ws.append(v)
                for cell in ws[ws_row_count:ws_row_count]:
                    if ws_row_count % 2 == 0: cell.style = 'odd'
                    else: cell.style = 'even'
                ws_row_count += 1
            
            # Save the Workbook
            wb.save(filename=workbook)


    #======================================================
    # Function - Server Inventory
    #======================================================
    def srv_inventory(self, kwargs):
        for k, v in kwargs.org_moids.items(): kwargs.org = k
        domains = DotMap()
        servers = DotMap()
        platform_types   = "('IMCBlade', 'IMCM4', 'IMCM5', 'IMCM6', 'IMCM7', 'IMCM8', 'IMCM9', 'UCSFI', 'UCSFIISM')"
        kwargs.api_filter= f"PlatformType in {platform_types}"
        kwargs.method    = 'get'
        kwargs.top1000   = True
        kwargs.qtype     = 'device_registration'
        kwargs.uri       = 'asset/DeviceRegistrations'
        kwargs           = isight.api(kwargs.qtype).calls(kwargs)
        for e in kwargs.results:
            if re.search('UCSFI(ISM)?', e.PlatformType):
                domains[e.Moid] = DotMap(
                    name = e.DeviceHostname[0],
                    serials = e.Serial,
                    servers = DotMap(),
                    type    = e.PlatformType)
        for e in kwargs.results:
            if re.search('IMC', e.PlatformType):
                parent = ''
                if e.get('ParentConnection'): parent = e.ParentConnection.Moid
                servers[e.Serial[0]] = DotMap(
                    server_name  = e.DeviceHostname[0],
                    parent       = parent,
                    registration = e.Moid,
                )
        kwargs.top1000 = True
        kwargs.qtype   = 'physical_summaries'
        kwargs.uri     = 'compute/PhysicalSummaries'
        kwargs         = isight.api(kwargs.qtype).calls(kwargs)
        for e in kwargs.results:
            servers[e.Serial].chassis_id     = e.ChassisId
            servers[e.Serial].hw_moid        = e.Moid
            servers[e.Serial].mgmt_ip_address= e.MgmtIpAddress
            servers[e.Serial].mgmt_mode      = e.ManagementMode
            servers[e.Serial].model          = e.Model
            servers[e.Serial].name           = e.Name
            servers[e.Serial].object_type    = e.SourceObjectType
            servers[e.Serial].organization   = DotMap()
            servers[e.Serial].platform       = e.PlatformType
            servers[e.Serial].power_state    = e.OperPowerState
            servers[e.Serial].registration   = e.RegisteredDevice.Moid
            servers[e.Serial].server_dn      = e.Dn
            servers[e.Serial].server_id      = e.ServerId
            servers[e.Serial].server_profile = e.ServiceProfile
            servers[e.Serial].slot           = e.SlotId
            servers[e.Serial].wwnn           = 'unassigned'
        # Obtain Server Profile Data
        kwargs.method = 'get'
        kwargs.top1000= True
        kwargs.qtype  = 'server'
        kwargs.uri    = 'server/Profiles'
        kwargs        = isight.api(kwargs.qtype).calls(kwargs)
        if kwargs.results == None: prRed('empty results.  Exiting script...')
        profile_moids = []
        temp_info = deepcopy(servers)

        for e in kwargs.results:
            org_name = [k for k, v in kwargs.org_moids.items() if v.moid == e.Organization.Moid][0]
            if e.get('AssociatedServer'):
                for b,c in temp_info.items():
                    if e.AssignedServer.Moid == c.hw_moid:
                        servers[b].server_profile = e.Name
                        servers[b].moid           = e.Moid
                        servers[b].organization   = DotMap(name = org_name, moid = kwargs.org_moids[org_name].moid)
                        profile_moids.append(e.Moid)
        names = "', '".join(profile_moids).strip("', '")
        kwargs.api_filter= f"PoolPurpose eq 'WWNN' and AssignedToEntity.Moid in ('{names}')"
        kwargs.build_skip= True
        kwargs.qtype     = 'wwnn_pool_leases'
        kwargs.uri       = 'fcpool/Leases'
        kwargs           = isight.api(kwargs.qtype).calls(kwargs)
        wwnn_leases      = kwargs.results
        kwargs.api_filter= f"Profile.Moid in ('{names}')"
        kwargs.qtype     = 'assigned_vnics'
        kwargs.uri       = 'vnic/FcIfs'
        kwargs           = isight.api(kwargs.qtype).calls(kwargs)
        vhbas            = kwargs.results
        kwargs.api_filter= f"Profile.Moid in ('{names}')"
        kwargs.qtype     = 'assigned_vnics'
        kwargs.uri       = 'vnic/EthIfs'
        kwargs           = isight.api(kwargs.qtype).calls(kwargs)
        vnics            = kwargs.results

        qos_names = []
        for e in vnics: qos_names.append(e.EthQosPolicy.Moid)
        names = "', '".join(qos_names).strip("', '")
        kwargs.api_filter= f"Moid in ('{names}')"
        kwargs.qtype     = 'qos_policies'
        kwargs.uri       = 'vnic/EthQosPolicies'
        kwargs           = isight.api(kwargs.qtype).calls(kwargs)
        qos_results      = kwargs.results
        qos_policies = DotMap()
        for e in qos_results:
            qos_policies[e.Moid].name = e.Name
            qos_policies[e.Moid].mtu  = e.Mtu
        temp_info = deepcopy(servers)
        for k, v in temp_info.items():
            servers[k].vhbas = []
            servers[k].vnics = []
            for e in wwnn_leases:
                if e.AssignedToEntity.Moid == v.moid: servers[k].wwnn = e.WwnId
            for e in vhbas:
                if e.Profile.Moid == v.moid:
                    if e.WwpnAddressType == 'STATIC':
                        servers[k].vhbas.append({'name': e.Name, 'switch_id':e.Placement.SwitchId,'wwpn_address':e.StaticWwpnAddress})
                    else: servers[k].vhbas.append({'name': e.Name, 'switch_id':e.Placement.SwitchId,'wwpn_address':e.Wwpn})
            #exit()
            for e in vnics:
                if e.Profile.Moid == v.moid:
                    servers[k].vnics.append({'name': e.Name, 'mac_address':e.MacAddress,'mtu':qos_policies[e.EthQosPolicy.Moid].mtu})
        for k, v in servers.items():
            servers[k].vnics.sort(key=lambda d: d['name'])
            servers[k].vhbas.sort(key=lambda d: d['name'])
            if v.platform == 'UCSFI': servers[k].domain = domains[v.registration].name
            elif len(v.parent) > 0: servers[k].domain = domains[v.parent].name

        if len(servers) > 0:
            servers = DotMap(sorted(servers.items()))
            kwargs = get_local_time(kwargs)

            # Build Named Style Sheets for Workbook
            if kwargs.args.full_inventory:
                workbook = f'UCS-Inventory-Collector-{kwargs.time_short}.xlsx'
                kwargs = workbook_styles(kwargs)
                wb = kwargs.wb
                ws = wb.active
                ws.title = 'Inventory List'
        
                # Read Server Inventory to Create Column Headers
                column_headers = ['Domain','Profile','Server','Serial']
                vhba_list = []; vnic_list = []
                for key, value in servers.items():
                    if value.get('wwnn'):
                        if not 'WWNN' in column_headers: column_headers.append('WWNN')
                    if value.get('vhbas'):
                        for e in value.vhbas:
                            if not e.name in vhba_list: vhba_list.append(e.name)
                    if value.get('vnics'):
                        for e in value.vnics:
                            if not e.name in vnic_list: vnic_list.append(e.name)
                vhba_list.sort();  vnic_list.sort()
                column_headers = column_headers + vhba_list + vnic_list
                for i in range(len(column_headers)): ws.column_dimensions[chr(ord('@')+i+1)].width = 30
                cLength = len(column_headers)
                ws_header = f'Collected UCS Data on {kwargs.time_long}'
                data = [ws_header]
                ws.append(data)
                ws.merge_cells(f'A1:{chr(ord("@")+cLength)}1')
                for cell in ws['1:1']: cell.style = 'heading_1'
                ws.append(column_headers)
                for cell in ws['2:2']: cell.style = 'heading_2'
                ws_row_count = 3
                
                # Populate the Columns with Server Inventory
                for k, v in servers.items():
                    data = []
                    for i in column_headers:
                        column_count = 0
                        if i == 'Domain':
                            if len(v.domain) == 0: data.append(''); column_count += 1
                            else: data.append(v.domain); column_count += 1
                        elif i == 'Profile':
                            if len(v.server_profile) == 0: data.append(''); column_count += 1
                            else: data.append(v.server_profile); column_count += 1
                        elif i == 'Server':
                            if not 'sys' in v.server_dn:
                                if len(v.chassis_id) > 0: server_dn = f'sys/chassis-{v.chassis_id}/blade-{v.slot}'
                                else: server_dn = f'sys/rack-unit-{v.server_id}'
                            else: server_dn = v.server_dn
                            if len(server_dn) == 0: data.append(''); column_count += 1
                            else: data.append(server_dn); column_count += 1
                        elif i == 'Serial':
                            if len(k) == 0: data.append(''); column_count += 1
                            else: data.append(k); column_count += 1
                        elif i == 'WWNN':
                            if len(v.wwnn) == 0: data.append(''); column_count += 1
                            else: data.append(v.wwnn); column_count += 1
                        else:
                            if v.get('vhbas'):
                                for e in v.vhbas:
                                    if i == e.name: data.append(e.wwpn_address); column_count += 1
                            if v.get('vnics'):
                                for e in v.vnics:
                                    if i == e.name: data.append(e.mac_address); column_count += 1
                        if column_count == 0: data.append('Not Configured')
                        
                    # Add the Columns to the Spreadsheet
                    ws.append(data)
                    for cell in ws[ws_row_count:ws_row_count]:
                        if ws_row_count % 2 == 0: cell.style = 'odd'
                        else: cell.style = 'even'
                    ws_row_count += 1
            else:
                workbook = f'UCS-WWPN-Collector-{kwargs.time_short}.xlsx'
                kwargs = workbook_styles(kwargs)
                wb = kwargs.wb
                ws = wb.active
                ws.title = 'WWPN List'
        
                # Read Server Inventory to Create Column Headers
                column_headers = ['Profile','Serial']
                vhba_list = []
                for key, value in servers.items():
                    if value.get('wwnn'):
                        if not 'WWNN' in column_headers: column_headers.append('WWNN')
                    if value.get('vhbas'):
                        for e in value.vhbas:
                            if not e.name in vhba_list: vhba_list.append(e.name)
                vhba_list.sort()
                column_headers= column_headers + vhba_list
                for i in range(len(column_headers)): ws.column_dimensions[chr(ord('@')+i+1)].width = 30
                cLength = len(column_headers)
                ws_header = f'Collected UCS Data on {kwargs.time_long}'
                data = [ws_header]
                ws.append(data)
                ws.merge_cells(f'A1:{chr(ord("@")+cLength)}1')
                for cell in ws['1:1']: cell.style = 'heading_1'
                ws.append(column_headers)
                for cell in ws['2:2']: cell.style = 'heading_2'
                ws_row_count = 3
                
                # Populate the Columns with Server Inventory
                for k, v in servers.items():
                    data = []
                    for i in column_headers:
                        column_count = 0
                        if i == 'Profile':
                            if len(v.server_profile) == 0: data.append(''); column_count += 1
                            else: data.append(v.server_profile); column_count += 1
                        elif i == 'Serial': data.append(k); column_count += 1
                        elif i == 'WWNN':
                            if len(v.wwnn) == 0: data.append(''); column_count += 1
                            else: data.append(v.wwnn); column_count += 1
                        else:
                            if v.get('vhbas'):
                                for e in v.vhbas:
                                    if i == e.name: data.append(e.wwpn_address); column_count += 1
                        if column_count == 0: data.append('Not Configured')
                        
                    # Add the Columns to the Spreadsheet
                    ws.append(data)
                    for cell in ws[ws_row_count:ws_row_count]:
                        if ws_row_count % 2 == 0: cell.style = 'odd'
                        else: cell.style = 'even'
                    ws_row_count += 1
                
            # Save the Workbook
            wb.save(filename=workbook)

#======================================================
# Function - Setup Local Time
#======================================================
def get_local_time(kwargs):
    timezones   = kwargs.ez_data.ntp.allOf[1].properties.timezone.enum
    tz_regions = list(set([e.split('/')[0] for e in timezones]))
    kwargs.jdata = DotMap(
        default      = sorted(tz_regions)[0],
        description  = f'Select the Timezone Region.',
        enum         = sorted(tz_regions),
        title        = f'{kwargs.qtype} Profiles')
    tz_region = ezfunctions.variablePrompt(kwargs)
    region_tzs = [e.split('/')[1] for e in timezones if tz_region in e]
    kwargs.jdata = DotMap(
        default      = sorted(region_tzs)[0],
        description  = f'Select the Timezone within the Region.',
        enum         = sorted(region_tzs),
        title        = f'{kwargs.qtype} Profiles')
    timezone = f'{tz_region}/{ezfunctions.variablePrompt(kwargs)}'
    kwargs.datetime = datetime.now(pytz.timezone(timezone))
    kwargs.time_short = kwargs.datetime.strftime('%Y-%m-%d-%H-%M')
    kwargs.time_long  = kwargs.datetime.strftime('%Y-%m-%d %H:%M:%S %Z %z')
    return kwargs


#======================================================
# Function - Prompt for Organizations
#======================================================
def select_organizations(kwargs):
    kwargs.jdata = DotMap(
        default      = 'default',
        description  = f'Select the Organizations to Apply the changes to.',
        enum         = kwargs.org_moids.keys(),
        multi_select = True,
        title        = 'Organizations')
    kwargs.organizations = ezfunctions.variablePrompt(kwargs)
    return kwargs


#======================================================
# Function - Update Profile
#======================================================
def update_profile(org, target, target_platform, yaml_arg, ydata, kwargs):
    kwargs.org = org
    policies = DotMap()
    for policy in target[target_platform].update_policies:
        #======================================================
        # Query API for List of Policies.
        #======================================================
        kwargs.api_filter= f"Organization.Moid eq '{kwargs.org_moids[org].moid}'"
        kwargs.method    = 'get'
        kwargs.qtype     = policy
        kwargs.uri       = kwargs.ez_data[policy].intersight_uri
        kwargs           = isight.api(policy).calls(kwargs)
        if kwargs.results == None: isight.empty_results(kwargs)
        policies[policy] = kwargs.pmoids
        #======================================================
        # Prompt User for Policy to Attach.
        #======================================================
        kwargs.jdata = DotMap(
            default      = list(policies[policy].keys())[0],
            description  = f'Select the {policy} Policy to attach to the Profile(s).',
            enum         = list(policies[policy].keys()),
            title        = f'{policy} Policy Name')
        policies[policy].name = ezfunctions.variablePrompt(kwargs)
    #======================================================
    # Obtain Profile Data.
    #======================================================
    kwargs.method  = 'get'
    kwargs.top1000 = True
    org_moid       = kwargs.org_moids[org].moid
    if re.search('^server(_template)?$', target_platform):
        target_type = kwargs.sub_type[target_platform]
        kwargs.api_filter= f"Organization.Moid eq '{org_moid}' and TargetPlatform eq '{target_type}'"
    else: kwargs.api_filter = f"Organization.Moid eq '{org_moid}'"
    kwargs.qtype      = target_platform
    kwargs.uri        = kwargs.ez_data[kwargs.qtype].intersight_uri
    kwargs = isight.api(target_platform).calls(kwargs)
    profiles = kwargs.pmoids
    for k,v in profiles.items(): profiles[k].name = k
    if profiles == None: isight.empty_results(kwargs)
    #======================================================
    # Request from User Which Profiles to Apply this to.
    #======================================================
    if yaml_arg == False:
        kwargs.jdata = DotMap(
            default      = sorted(list(profiles.keys()))[0],
            description  = f'Select the {kwargs.qtype} Profiles to Apply the {policy} to.',
            enum         = sorted(list(profiles.keys())),
            multi_select = True,
            title        = f'{kwargs.qtype} Profiles')
        profile_names = ezfunctions.variablePrompt(kwargs)
    else: profile_names = ydata.profiles
    #======================================================
    # Attach the Policy to the Selected Server Profiles.
    #======================================================
    for i in profile_names:
        prPurple(f'\n{"-"*91}\n  Starting on Server Profile `{i}`.\n{"-"*91}\n')
        policy_bucket = profiles[i].policy_bucket
        object_index  = dict((d.ObjectType, s) for s, d in enumerate(policy_bucket))
        for e in policies.keys():
            policy_moid = policies[e][policies[e].name].moid
            object_type = kwargs.ez_data[e].ObjectType
            policy_uri  = kwargs.ez_data[e].intersight_uri
            #======================================================
            # Index the Server List to find the Server Profile and 
            # pull the Policy BucketSee if the Policy Type is 
            # Already Attached.  If attached, Update to the new.
            # Moid, else attach the Policy.
            #======================================================
            if object_index.get(object_type):
                type_index = object_index.get(object_type, -1)
                policy_link = f"https://www.intersight.com/api/v1/{policy_uri}/{policy_moid}"
                policy_bucket[type_index].update({'Moid':policy_moid, 'link': policy_link})
            else:
                policy_bucket.append({'Moid': policy_moid, 'ObjectType': object_type})
        pbucket = []
        for x in policy_bucket: pbucket.append(x.toDict())
        policy_bucket = (sorted(pbucket, key=lambda k: (k['ObjectType'])))
        #======================================================
        # Patch the Profile with new Policy Bucket.
        #======================================================
        kwargs.apiBody = {"Name":profiles[i].name,"PolicyBucket":pbucket}
        kwargs.method = 'patch'
        kwargs.pmoid  = profiles[i].moid
        if re.search('FIAttached|Standalone', target_platform):
            kwargs.uri   = kwargs.ez_data['server'].intersight_uri
        else: kwargs.uri = kwargs.ez_data[target_platform].intersight_uri
        kwargs = isight.api(target_platform).calls(kwargs)


#======================================================
# Function - Workbook Styles
#======================================================
def workbook_styles(kwargs):
    wb = openpyxl.Workbook()
    # Build Named Style Sheets for Workbook
    bd1 = Side(style="thick", color="0070C0")
    bd2 = Side(style="medium", color="0070C0")
    heading_1 = NamedStyle(name="heading_1")
    heading_1.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    heading_1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
    heading_1.fill = PatternFill("solid", fgColor="305496")
    heading_1.font = Font(bold=True, size=15, color="FFFFFF")
    heading_2 = NamedStyle(name="heading_2")
    heading_2.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    heading_2.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    heading_2.font = Font(bold=True, size=15, color="44546A")
    even = NamedStyle(name="even")
    even.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    even.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
    even.font = Font(bold=False, size=12, color="44546A")
    odd = NamedStyle(name="odd")
    odd.alignment = Alignment(horizontal="center", vertical="center", wrap_text="True")
    odd.border = Border(left=bd2, top=bd2, right=bd2, bottom=bd2)
    odd.fill = PatternFill("solid", fgColor="D9E1F2")
    odd.font = Font(bold=False, size=12, color="44546A")
    wb.add_named_style(heading_1)
    wb.add_named_style(heading_2)
    wb.add_named_style(even)
    wb.add_named_style(odd)
    kwargs.wb = wb
    return kwargs
