from classes import validatingv2 as validating
from dotmap import DotMap
from copy import deepcopy
from git import cmd, Repo
from openpyxl import load_workbook
from ordered_set import OrderedSet
import ipaddress
import itertools
import jinja2
import json
import os
import pexpect
import pkg_resources
import re
import requests
import shutil
import subprocess
import sys
import stdiomask
import textwrap
import yaml

# Log levels 0 = None, 1 = Class only, 2 = Line
log_level = 2

# Exception Classes
class InsufficientArgs(Exception):
    pass

class ErrException(Exception):
    pass

class InvalidArg(Exception):
    pass

class LoginFailed(Exception):
    pass

class MyDumper(yaml.Dumper):
    def increase_indent(self, flow=False, indentless=False):
        return super(MyDumper, self).increase_indent(flow, False)

#=====================================================
# Print Color Functions
#=====================================================
def prCyan(skk): print("\033[96m {}\033[00m" .format(skk))
def prGreen(skk): print("\033[92m {}\033[00m" .format(skk))
def prLightPurple(skk): print("\033[94m {}\033[00m" .format(skk))
def prLightGray(skk): print("\033[97m {}\033[00m" .format(skk))
def prPurple(skk): print("\033[95m {}\033[00m" .format(skk))
def prRed(skk): print("\033[91m {}\033[00m" .format(skk))
def prYellow(skk): print("\033[93m {}\033[00m" .format(skk))

#=====================================================
# pexpect - Login Function
#=====================================================
def child_login(kwargs):
    kwargs.sensitive_var = kwargs.password
    kwargs   = sensitive_var_value(kwargs)
    password = kwargs.var_value
    #=====================================================
    # Use 
    #=====================================================
    if kwargs.op_system == 'Windows':
        from pexpect import popen_spawn
        child = popen_spawn.PopenSpawn('cmd', encoding='utf-8', timeout=1)
    else:
        system_shell = os.environ['SHELL']
        child = pexpect.spawn(system_shell, encoding='utf-8')
    child.logfile_read = sys.stdout
    if kwargs.op_system == 'Windows':
        child.sendline(f'ping -n 2 {kwargs.hostname}')
        child.expect(f'ping -n 2 {kwargs.hostname}')
        child.expect_exact("> ")
        child.sendline(f'ssh {kwargs.username}@{kwargs.hostname} | Tee-Object {kwargs.hostname}.txt')
        child.expect(f'Tee-Object {kwargs.hostname}.txt')
    else:
        child.sendline(f'ping -c 2 {kwargs.hostname}')
        child.expect(f'ping -c 2 {kwargs.hostname}')
        child.expect_exact("$ ")
        child.sendline(f'ssh {kwargs.username}@{kwargs.hostname} | tee {kwargs.hostname}.txt')
        child.expect(f'tee {kwargs.hostname}.txt')
    logged_in = False
    while logged_in == False:
        i = child.expect(['Are you sure you want to continue', 'closed', 'Password:', kwargs.host_prompt, pexpect.TIMEOUT])
        if i == 0: child.sendline('yes')
        elif i == 1:
            prRed(f'\n!!! FAILED !!! to connect.  '\
                f'Please Validate {kwargs.hostname} is correct and username {kwargs.username} is correct.')
            sys.exit(1)
        elif i == 2: child.sendline(password)
        elif i == 3: logged_in = True
        elif i == 4:
            prRed(f"\n{'-'*91}\n")
            prRed(f'!!! FAILED !!!\n Could not open SSH Connection to {kwargs.hostname}')
            prRed(f"\n{'-'*91}\n")
            sys.exit(1)
    return child, kwargs

#======================================================
# Function - Format Policy Description
#======================================================
def choose_policy(policy_type, kwargs):
    policy_descr = mod_pol_description(policy_type)
    policy_list = []
    for i in kwargs['policies'][policy_type]: policy_list.append(i['name'])
    valid = False
    while valid == False:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        if kwargs.get('optional_message'): print(kwargs['optional_message'])
        print(f'  {policy_descr} Policy Options:')
        for i, v in enumerate(policy_list):
            i += 1
            if i < 10: print(f'     {i}. {v}')
            else: print(f'    {i}. {v}')
        if kwargs['allow_opt_out'] == True: print(f'     99. Do not assign a(n) {policy_descr}.')
        print(f'     100. Create a New {policy_descr}.')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        policyOption = input(f"Select the Option Number for the {policy_descr} Policy to Assign to {kwargs['name']} Policy: ")
        if re.search(r'^[0-9]{1,3}$', policyOption):
            for i, v in enumerate(policy_list):
                i += 1
                if int(policyOption) == i:
                    kwargs['policy'] = v
                    valid = True
                    return kwargs
                elif int(policyOption) == 99:
                    kwargs['policy'] = ''
                    valid = True
                    return kwargs
                elif int(policyOption) == 100:
                    kwargs['policy'] = 'create_policy'
                    valid = True
                    return kwargs
            if int(policyOption) == 99:
                kwargs['policy'] = ''
                valid = True
                return kwargs
            elif int(policyOption) == 100:
                kwargs['policy'] = 'create_policy'
                valid = True
                return kwargs
        else: message_invalid_selection()

#======================================================
# Function - Count the Number of Keys
#======================================================
def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func:
                count += 1
    return count

#==========================================================
# Function for Processing easyDict and Creating YAML Files
#==========================================================
def create_yaml(orgs, kwargs):
    baseRepo = kwargs.args.dir
    ez_data   = kwargs.ez_data.ezimm.allOf[1].properties
    classes  = ez_data.classes.enum

    def write_file(dest_dir, dest_file, dict, title1):
        if not os.path.exists(os.path.join(dest_dir, dest_file)):
            create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'
            os.system(create_file)
        wr_file = open(os.path.join(dest_dir, dest_file), 'w')
        wr_file.write('---\n')
        wr_file = open(os.path.join(dest_dir, dest_file), 'a')
        dash_length = '='*(len(title1) + 20)
        wr_file.write(f'#{dash_length}\n')
        wr_file.write(f'#   {title1} - Variables\n')
        wr_file.write(f'#{dash_length}\n')
        wr_file.write(yaml.dump(dict, Dumper=MyDumper, default_flow_style=False))
        wr_file.close()
    for item in classes:
        dest_dir = os.path.join(baseRepo, ez_data[f'class.{item}'].directory)
        if item == 'policies':
            if not os.path.isdir(dest_dir): os.makedirs(dest_dir)
            for i in ez_data[f'class.{item}'].enum:
                idict = DotMap()
                print(orgs)
                for org in orgs:
                    if not idict.get(org):
                        idict[org] = DotMap()
                    for x in ez_data[f'class.{i}'].enum:
                        if kwargs.imm_dict.orgs[org].get(item):
                            if kwargs.imm_dict.orgs[org][item].get(x):
                                idict[org][item][x] = deepcopy(kwargs.imm_dict.orgs[org][item][x])
                            if not len(idict[org][item][x]) > 0: idict[org][item].pop(x)
                if len(idict[org][item]) > 0:
                    idict = json.dumps(idict.toDict())
                    idict = json.loads(idict)
                    for x in ez_data[f'class.{i}'].enum:
                        if kwargs.imm_dict.orgs[org][item].get(x):
                            if type(idict[org][item][x]) == list:
                                if idict[org][item][x][0].get('name'):
                                    idict[org][item][x] = list({v['name']:v for v in idict[org][item][x]}.values())
                                elif idict[org][item][x][0].get('names'):
                                    idict[org][item][x] = list({v['names'][0]:v for v in idict[org][item][x]}.values())
                    dest_file = f"{i}.yaml"
                    title1 = f"{str.title(item)} -> {i}"
                    write_file(dest_dir, dest_file, idict, title1)
        else:
            if not os.path.isdir(dest_dir): os.makedirs(dest_dir)
            for i in ez_data[f'class.{item}'].enum:
                idict = deepcopy(DotMap())
                if item == i:
                    for org in orgs:
                        if kwargs.imm_dict.orgs[org].get(item):
                            if len(kwargs.imm_dict.orgs[org][item]) > 0:
                                itemDict = deepcopy(kwargs.imm_dict.orgs[org][item].toDict())
                                idict[org][item] = itemDict
                                idict = json.dumps(idict.toDict())
                                idict = json.loads(idict)
                                if type(idict[org][item]) == list:
                                    idict[org][item] = list({v['name']:v for v in value}.values())
                                else:
                                    newdict = deepcopy(idict)
                                    if re.search('(netapp|storage)', item):
                                        for key, value in newdict[org][item].items():
                                            idict[org][item][key] = list({v['name']:v for v in value}.values())
                                    else:
                                        for key, value in newdict[org][item].items():
                                            idict[org][item][key] = list({v['name']:v for v in value}.values())
                                dest_file = f'{i}.yaml'
                                title1 = str.title(item.replace('_', ' '))
                                write_file(dest_dir, dest_file, idict, title1)
                else:
                    for org in orgs:
                        if kwargs.imm_dict.orgs[org].get(item):
                            if kwargs.imm_dict.orgs[org][item].get(i):
                                if len(kwargs.imm_dict.orgs[org][item][i]) > 0:
                                    idict[org][item][i] = deepcopy(kwargs.imm_dict.orgs[org][item][i])
                                    idict = json.dumps(idict.toDict())
                                    idict = json.loads(idict)
                                    if type(idict[org][item][i]) == list:
                                        if re.search('(chassis|server)', i) and item == 'profiles':
                                            idict[org][item][i] = list({
                                                v['targets'][0]['name']:v for v in idict[org][item][i]}.values())
                                        else:
                                            idict[org][item][i] = list(
                                                {v['name']:v for v in idict[org][item][i]}.values())
                                    else:
                                        for a, b in idict[org][item][i].items():
                                            b = list({v['name']:v for v in b}.values())
                                    dest_file = f'{i}.yaml'
                                    title1 = f"{str.title(item.replace('_', ' '))} -> {str.title(i.replace('_', ' '))}"
                                    write_file(dest_dir, dest_file, idict, title1)
                        
#======================================================
# Function - Ask User to Configure Additional Policy
#======================================================
def exit_default(policy_type, y_or_n):
    valid_exit = False
    while valid_exit == False:
        exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [{y_or_n}]: ')
        if exit_answer == '': exit_answer = y_or_n
        if exit_answer == 'N':
            policy_loop = True
            configure_loop = True
            valid_exit = True
        elif exit_answer == 'Y':
            policy_loop = False
            configure_loop = False
            valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, policy_loop

#======================================================
# Function - Ask User to Configure Additional Policy
#======================================================
def exit_default_del_tfc(policy_type, y_or_n):
    valid_exit = False
    while valid_exit == False:
        exit_answer = input(f'Would You like to {policy_type}?  Enter "Y" or "N" [{y_or_n}]: ')
        if exit_answer == '': exit_answer = y_or_n
        if exit_answer == 'N':
            policy_loop = True
            configure_loop = True
            valid_exit = True
        elif exit_answer == 'Y':
            policy_loop = False
            configure_loop = False
            valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, policy_loop

#======================================================
# Function - Prompt User with question
#======================================================
def exit_loop_default_yes(loop_count, policy_type):
    valid_exit = False
    while valid_exit == False:
        if loop_count % 2 == 0:
            exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [Y]: ')
        else: exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [N]: ')
        if (loop_count % 2 == 0 and exit_answer == '') or exit_answer == 'Y':
            configure_loop = False
            loop_count += 1
            policy_loop = False
            valid_exit = True
        elif not loop_count % 2 == 0 and exit_answer == '':
            configure_loop = True
            loop_count += 1
            policy_loop = True
            valid_exit = True
        elif exit_answer == 'N':
            configure_loop = True
            loop_count += 1
            policy_loop = True
            valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, loop_count, policy_loop

#========================================================
# Function to Append the imm_dict Dictionary
#========================================================
def ez_append(polVars, kwargs):
    class_path= kwargs['class_path']
    p         = class_path.split(',')
    org       = kwargs['org']
    polVars   = ez_remove_empty(polVars)
    polVars   = DotMap(polVars)
    # Confirm the Key Exists
    if not kwargs.imm_dict.orgs.get(org):
        kwargs.imm_dict.orgs[org] = DotMap()
    if len(p) >= 2:
        if not kwargs.imm_dict.orgs[org].get(p[0]):
            kwargs.imm_dict.orgs[org][p[0]] = DotMap()
    if len(p) >= 3:
        if not kwargs.imm_dict.orgs[org][p[0]].get(p[1]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]] = DotMap()
    if len(p) >= 4:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]].get(p[2]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]] = DotMap()
    if len(p) == 2:
        if not kwargs.imm_dict.orgs[org][p[0]].get(p[1]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]] = [deepcopy(polVars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]].append(deepcopy(polVars))
    elif len(p) == 3:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]].get(p[2]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]] = [deepcopy(polVars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].append(deepcopy(polVars))
    elif len(p) == 4:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].get(p[3]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]] = [deepcopy(polVars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]].append(deepcopy(polVars))
    elif len(p) == 5:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].get(p[3]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]] = DotMap()
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]].get(p[4]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]][p[4]] = [deepcopy(polVars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]][p[4]].append(deepcopy(polVars))

    return kwargs

#========================================================
# Function to Append the imm_dict Dictionary
#========================================================
def ez_append_wizard(polVars, kwargs):
    class_path = kwargs['class_path']
    p = class_path.split(',')
    polVars = ez_remove_empty(polVars)
    # Confirm the Key Exists
    if not kwargs.imm_dict.get('wizard'): kwargs.imm_dict['wizard'] = {}
    if len(p) >= 2:
        if not kwargs.imm_dict['wizard'].get(p[0]):
            kwargs.imm_dict['wizard'].update(deepcopy({p[0]:{}}))
    if len(p) >= 3:
        if not kwargs.imm_dict['wizard'][p[0]].get(p[1]):
            kwargs.imm_dict['wizard'][p[0]].update(deepcopy({p[1]:{}}))
    if len(p) == 1:
        if not kwargs.imm_dict['wizard'].get(p[0]):
            kwargs.imm_dict['wizard'].update(deepcopy({p[0]:[]}))
    elif len(p) == 2:
        if not kwargs.imm_dict['wizard'][p[0]].get(p[1]):
            kwargs.imm_dict['wizard'][p[0]].update(deepcopy({p[1]:[]}))
    elif len(p) == 3:
        if not kwargs.imm_dict['wizard'][p[0]][p[1]].get(p[2]):
            kwargs.imm_dict['wizard'][p[0]][p[1]].update(deepcopy({p[2]:[]}))
    # append the Dictionary
    if len(p) == 1: kwargs.imm_dict['wizard'][p[0]].append(deepcopy(polVars))
    if len(p) == 2: kwargs.imm_dict['wizard'][p[0]][p[1]].append(deepcopy(polVars))
    elif len(p) == 3: kwargs.imm_dict['wizard'][p[0]][p[1]][p[2]].append(deepcopy(polVars))
    return kwargs

#========================================================
# Function to Remove Empty Arguments
#========================================================
def ez_remove_empty(polVars):
    pop_list = []
    for k,v in polVars.items():
        if v == None:
            pop_list.append(k)
    for i in pop_list:
        polVars.pop(i)
    return polVars

#======================================================
# Function - find the Keys for each Section
#======================================================
def findKeys(ws, func_regex):
    func_list = OrderedSet()
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)):
                func_list.add(str(i[0].value))
    return func_list

#======================================================
# Function - Assign the Variables to the Keys
#======================================================
def findVars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            try:
                for x in range(2, 34):
                    if (ws.cell(row=i - 1, column=x)).value:
                        var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else:
                        x += 1
            except Exception as e:
                e = e
                pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list:
            var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value
            var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

#========================================================
# Function - Prompt User for the Intersight Configurtion
#========================================================
def intersight_config(kwargs):
    kwargs.jData = DotMap()
    if kwargs.args.api_key_id == None:
        kwargs.sensitive_var = 'intersight_apikey'
        kwargs = sensitive_var_value(kwargs)
        kwargs.args.api_key_id = kwargs.var_value

    #==============================================
    # Prompt User for Intersight SecretKey File
    #==============================================
    secret_path = kwargs.args.api_key_file
    secret_loop = False
    while secret_loop == False:
        valid = False
        if secret_path == None:
            varName = 'intersight_keyfile'
            print(f"\n-------------------------------------------------------------------------------------------\n")
            print(f"  The Script did not find {varName} as an 'environment' variable.")
            print(f"  To not be prompted for the value of {varName} each time")
            print(f"  add the following to your local environemnt:\n")
            print(f"    - Linux: export {varName}='{varName}_value'")
            print(f"    - Windows: $env:{varName}='{varName}_value'")
            secret_path = ''
        if '~' in secret_path: secret_path = os.path.expanduser(secret_path)
        if not secret_path == '':
            if not os.path.isfile(secret_path):
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  !!!Error!!! intersight_keyfile not found.')
            else:
                secret_file = open(secret_path, 'r')
                count = 0
                if '-----BEGIN RSA PRIVATE KEY-----' in secret_file.read(): count += 1
                secret_file.seek(0)
                if '-----END RSA PRIVATE KEY-----' in secret_file.read(): count += 1
                if count == 2:
                    kwargs.args.api_key_file = secret_path
                    secret_loop = True
                    valid = True
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'  !!!Error!!! intersight_keyfile does not seem to contain a Valid Secret Key.')
        if not valid == True:
            kwargs.jData.description = 'Intersight SecretKey'
            kwargs.jData.default = '%s%sDownloads%sSecretKey.txt' % (kwargs['home'], os.sep, os.sep)
            kwargs.jData.pattern = '.*'
            kwargs.jData.varInput= 'What is the Path for the Intersight SecretKey?'
            kwargs.jData.varName = 'intersight_keyfile'
            secret_path = varStringLoop(kwargs)

    #==============================================
    # Prompt User for Intersight Endpoint
    #==============================================
    valid = False
    while valid == False:
        varValue = kwargs['args'].endpoint
        if not varValue == None:
            varName = 'Intersight Endpoint'
            if re.search(r'^[a-zA-Z0-9]:', varValue):
                valid = validating.ip_address(varName, varValue)
            if re.search(r'[a-zA-Z]', varValue):
                valid = validating.dns_name(varName, varValue)
            elif re.search(r'^([0-9]{1,3}\.){3}[0-9]{1,3}$', varValue):
                valid = validating.ip_address(varName, varValue)
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print('  "{}" is not a valid address.').format(varValue)
                print(f'\n-------------------------------------------------------------------------------------------\n')
        if valid == False:
            kwargs.jData.description = 'Hostname of the Intersight Endpoint'
            kwargs.jData.default = 'intersight.com'
            kwargs.jData.pattern = '.*'
            kwargs.jData.varInput= 'What is the Intersight Endpoint Hostname?'
            kwargs.jData.varName = 'Intersight Endpoint'
            kwargs.jData.varType = 'hostname'
            kwargs.args.endpoint = varStringLoop(kwargs)
            valid = True
    # Return kwargs
    return kwargs

#======================================================
# Function - Print with json dumps
#======================================================
def jprint(jDict):
    print(json.dumps(jDict, indent=4))

#======================================================
# Function - Load Previous YAML Files
#======================================================
def load_previous_configurations(kwargs):
    ez_data   = kwargs.ez_data.ezimm.allOf[1].properties
    vclasses = ez_data.classes.enum
    dir_check   = 0
    if os.path.isdir(kwargs.args.dir):
        dir_list = os.listdir(kwargs.args.dir)
        for i in dir_list:
            if i == 'policies': dir_check += 1
            elif i == 'pools': dir_check += 1
            elif i == 'profiles': dir_check += 1
            elif i == 'templates': dir_check += 1
    if dir_check > 1:
        for item in vclasses:
            dest_dir = ez_data[f'class.{item}'].directory
            if os.path.isdir(os.path.join(kwargs.args.dir, dest_dir)):
                dir_list = os.listdir(os.path.join(kwargs.args.dir, dest_dir))
                for i in dir_list:
                    yfile = open(os.path.join(kwargs.args.dir, dest_dir, i), 'r')
                    data = yaml.safe_load(yfile)
                    for key, value in data.items():
                        if not kwargs.imm_dict.orgs.get(key): kwargs.imm_dict.orgs[key] = {}
                        for k, v in value.items():
                            if not kwargs.imm_dict.orgs[key].get(k): kwargs.imm_dict.orgs[key][k] = {}
                            kwargs.imm_dict.orgs[key][k].update(deepcopy(v))
    # Return kwargs
    return kwargs

#======================================================
# Function - Local User Policy - Users
#======================================================
def local_users_function(kwargs):
    ez_data = kwargs.ez_data
    inner_loop_count = 1
    json_data = kwargs['json_data']
    kwargs['local_users'] = []
    valid_users = False
    while valid_users == False:
        kwargs['multi_select'] = False
        jsonVars = json_data['iam.EndPointUser'].allOf[1].properties
        #==============================================
        # Prompt User for Local Username
        #==============================================
        kwargs['jData'] = deepcopy(jsonVars['Name'])
        kwargs['jData']['default'] = 'admin'
        kwargs['jData']['minimum'] = 1
        kwargs['jData']['varInput'] = 'What is the Local username?'
        kwargs['jData']['varName'] = 'Local User'
        username = varStringLoop(kwargs)
        #==============================================
        # Prompt User for User Role
        #==============================================
        jsonVars = ez_data.ezimm.allOf[1].properties['policies']['iam.LocalUserPasswordPolicy']
        kwargs['jData'] = deepcopy(jsonVars['role'])
        kwargs['jData']['varType'] = 'User Role'
        role = variablesFromAPI(kwargs)

        if kwargs['enforce_strong_password'] == True:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print('Enforce Strong Password is enabled.  The following rules must be followed:')
            print('  - The password must have a minimum of 8 and a maximum of 20 characters.')
            print("  - The password must not contain the User's Name.")
            print('  - The password must contain characters from three of the following four categories.')
            print('    * English uppercase characters (A through Z).')
            print('    * English lowercase characters (a through z).')
            print('    * Base 10 digits (0 through 9).')
            print('    * Non-alphabetic characters (! , @, #, $, %, ^, &, *, -, _, +, =)\n\n')
        kwargs.sensitive_var = f'local_user_password_{inner_loop_count}'
        kwargs = sensitive_var_value(kwargs)
        user_attributes = {
            'enabled':True,
            'password':inner_loop_count,
            'role':role,
            'username':username
        }
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'   enabled  = True')
        print(f'   password = "Sensitive"')
        print(f'   role     = "{role}"')
        print(f'   username = "{username}"')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        valid_confirm = False
        while valid_confirm == False:
            question = input('Do you want to accept the above configuration?  Enter "Y" or "N" [Y]: ')
            if question == 'Y' or question == '':
                kwargs['local_users'].append(user_attributes)
                valid_exit = False
                while valid_exit == False:
                    loop_exit = input(f'Would You like to Configure another Local User?  Enter "Y" or "N" [N]: ')
                    if loop_exit == 'Y':
                        inner_loop_count += 1
                        valid_confirm = True
                        valid_exit = True
                    elif loop_exit == 'N' or loop_exit == '':
                        valid_confirm = True
                        valid_exit = True
                        valid_users = True
                    else: message_invalid_y_or_n('short')
            elif question == 'N':
                policy_type = 'Local User'
                message_starting_over(policy_type)
                valid_confirm = True
            else: message_invalid_y_or_n('long')
    return kwargs

#======================================================
# Function - Merge Easy IMM Repository to Dest Folder
#======================================================
def merge_easy_imm_repository(orgs, kwargs):
    baseRepo       = kwargs['args'].dir
    # Download the Easy IMM Comprehensive Example Base Repo
    tfe_dir = 'tfe_modules'
    git_url = "https://github.com/terraform-cisco-modules/easy-imm-comprehensive-example"
    if not os.path.isdir(tfe_dir):
        os.mkdir(tfe_dir)
        Repo.clone_from(git_url, tfe_dir)
    if not os.path.isfile(os.path.join(tfe_dir, 'README.md')): Repo.clone_from(git_url, tfe_dir)
    else: g = cmd.Git(tfe_dir); g.pull()
    if not os.path.isdir(baseRepo): os.mkdir(baseRepo)
    # Now Loop over the folders and merge the module files
    for folder in ['defaults', '']:
        if folder == 'defaults':
            dest_dir = os.path.join(baseRepo, folder)
            src_dir = os.path.join(tfe_dir, 'defaults')
            if not os.path.isdir(dest_dir): os.mkdir(dest_dir)
        else:
            dest_dir = os.path.join(baseRepo)
            src_dir = os.path.join(tfe_dir)
        copy_files = os.listdir(src_dir)
        for fname in copy_files:
            if not os.path.isdir(os.path.join(src_dir, fname)):
                shutil.copy2(os.path.join(src_dir, fname), dest_dir)

#======================================================
# Function - Message for Invalid List Selection
#======================================================
def message_invalid_selection():
    print(f'\n-------------------------------------------------------------------------------------------\n')
    print(f'  Error!! Invalid Selection.  Please Select a valid Option from the List.')
    print(f'\n-------------------------------------------------------------------------------------------\n')

#======================================================
# Function - Message for Invalid Selection Y or N
#======================================================
def message_invalid_y_or_n(length):
    if length == 'short': dashRep = '-'*54
    else: dashRep = '-'*91
    print(f'\n{dashRep}\n')
    print(f'  Error!! Invalid Value.  Please enter "Y" or "N".')
    print(f'\n{dashRep}\n')

#======================================================
# Function - Message Invalid FCoE VLAN
#======================================================
def message_fcoe_vlan(fcoe_id, vlan_policy):
    dashRep = '-'*91
    print(f'\n{dashRep}\n')
    print(f'  Error!!  The FCoE VLAN {fcoe_id} is already assigned to the VLAN Policy')
    print(f'  {vlan_policy}.  Please choose a VLAN id that is not already in use.')
    print(f'\n{dashRep}\n')

#======================================================
# Function - Message Invalid Native VLAN
#======================================================
def message_invalid_native_vlan(nativeVlan, VlanList):
    dashRep = '-'*91
    print(f'\n{dashRep}\n')
    print(f'  Error!! The Native VLAN "{nativeVlan}" was not in the VLAN Policy List.')
    print(f'  VLAN Policy List is: "{VlanList}"')
    print(f'\n{dashRep}\n')

#======================================================
# Function - Message Invalid VLAN
#======================================================
def message_invalid_vlan():
    dashRep = '-'*91
    print(f'\n{dashRep}\n')
    print(f'  Invalid Entry!  Please Enter a valid VLAN ID in the range of 1-4094.')
    print(f'\n{dashRep}\n')

#======================================================
# Function - Message Invalid VSAN
#======================================================
def message_invalid_vsan():
    dashRep = '-'*91
    print(f'\n{dashRep}\n')
    print(f'  Invalid Entry!  Please Enter a valid VSAN ID in the range of 1-4094.')
    print(f'\n{dashRep}\n')

#======================================================
# Function - Message Invalid VLAN
#======================================================
def message_invalid_vsan_id(vsan_policy, vsan_id, vsan_list):
    print(f'\n-------------------------------------------------------------------------------------------\n')
    print(f'  Error with VSAN!!  The VSAN {vsan_id} is not in the VSAN Policy')
    print(f'  {vsan_policy}.  Options are {vsan_list}.')
    print(f'\n-------------------------------------------------------------------------------------------\n')

#======================================================
# Function - Message Starting Over
#======================================================
def message_starting_over(policy_type):
    dashRep = '-'*54
    print(f'\n{dashRep}\n')
    print(f'  Starting {policy_type} Section over.')
    print(f'\n{dashRep}\n')

#======================================================
# Function - Change Policy Description to Sentence
#======================================================
def mod_pol_description(pol_description):
    pol_description = str.title(pol_description.replace('_', ' '))
    pol_description = pol_description.replace('Ipmi', 'IPMI')
    pol_description = pol_description.replace('Ip', 'IP')
    pol_description = pol_description.replace('Iqn', 'IQN')
    pol_description = pol_description.replace('Ldap', 'LDAP')
    pol_description = pol_description.replace('Ntp', 'NTP')
    pol_description = pol_description.replace('Sd', 'SD')
    pol_description = pol_description.replace('Smtp', 'SMTP')
    pol_description = pol_description.replace('Snmp', 'SNMP')
    pol_description = pol_description.replace('Ssh', 'SSH')
    pol_description = pol_description.replace('Wwnn', 'WWNN')
    pol_description = pol_description.replace('Wwnn', 'WWNN')
    pol_description = pol_description.replace('Wwpn', 'WWPN')
    pol_description = pol_description.replace('Vsan', 'VSAN')
    return pol_description

#======================================================
# Function - Naming Rule
#======================================================
def naming_rule(name_prefix, name_suffix, org):
    if not name_prefix == '':
        name = '%s_%s' % (name_prefix, name_suffix)
    else:
        name = '%s_%s' % (org, name_suffix)
    return name

#======================================================
# Function - Naming Rule Fabric Policy
#======================================================
def naming_rule_fabric(loop_count, name_prefix, org):
    if loop_count % 2 == 0:
        if not name_prefix == '': name = f'{name_prefix}-a'
        elif not org == 'default': name = f'{org}-a'
        else: name = 'fabric-a'
    else:
        if not name_prefix == '': name = f'{name_prefix}-b'
        elif not org == 'default': name = f'{org}-a'
        else: name = 'fabric-b'
    return name

#======================================================
# Function - NTP
#======================================================
def ntp_alternate():
    valid = False
    while valid == False:
        alternate_true = input('Do you want to Configure an Alternate NTP Server?  Enter "Y" or "N" [Y]: ')
        if alternate_true == 'Y' or alternate_true == '':
            alternate_ntp = input('What is your Alternate NTP Server? [1.north-america.pool.ntp.org]: ')
            if alternate_ntp == '': alternate_ntp = '1.north-america.pool.ntp.org'
            if re.search(r'[a-zA-Z]+', alternate_ntp): valid = validating.dns_name('Alternate NTP Server', alternate_ntp)
            else: valid = validating.ip_address('Alternate NTP Server', alternate_ntp)
        elif alternate_true == 'N':
            alternate_ntp = ''
            valid = True
        else: message_invalid_y_or_n('short')
    return alternate_ntp

#======================================================
# Function - NTP
#======================================================
def ntp_primary():
    valid = False
    while valid == False:
        primary_ntp = input('What is your Primary NTP Server [0.north-america.pool.ntp.org]: ')
        if primary_ntp == "": primary_ntp = '0.north-america.pool.ntp.org'
        if re.search(r'[a-zA-Z]+', primary_ntp): valid = validating.dns_name('Primary NTP Server', primary_ntp)
        else: valid = validating.ip_address('Primary NTP Server', primary_ntp)
    return primary_ntp

#======================================================
# Function - Get Policies from Dictionary
#======================================================
def policies_parse(ptype, policy_type, kwargs):
    org  = kwargs['org']
    kwargs['policies'] = []
    if not kwargs.imm_dict.orgs[org].get(ptype) == None:
        if not kwargs.imm_dict.orgs[org][ptype].get(policy_type) == None:
            kwargs['policies'] = {policy_type:kwargs.imm_dict.orgs[org][ptype][policy_type]}
        else: kwargs['policies'] = {policy_type:{}}
    else: kwargs['policies'] = {policy_type:{}}
    return kwargs

#======================================================
# Function - Prompt User to Enter Policy Description
#======================================================
def policy_descr(name, policy_type):
    valid = False
    while valid == False:
        descr = input(f'What is the Description for the {policy_type}?  [{name} {policy_type}]: ')
        if descr == '': descr = '%s %s' % (name, policy_type)
        valid = validating.description(f"{policy_type} polVars['descr']", descr, 1, 62)
        if valid == True: return descr

#======================================================
# Function - Prompt User to Enter Policy Name
#======================================================
def policy_name(namex, policy_type):
    valid = False
    while valid == False:
        name = input(f'What is the Name for the {policy_type}?  [{namex}]: ')
        if name == '': name = '%s' % (namex)
        valid = validating.name_rule(f'{policy_type} Name', name, 1, 62)
        if valid == True: return name

#======================================================
# Function - Validate input for each method
#======================================================
def process_kwargs(kwargs):
    # Validate User Input
    json_data = kwargs['validateData']
    validate_args(kwargs)

    error_count = 0
    error_list = []
    optional_args = json_data['optional_args']
    required_args = json_data['required_args']
    for item in required_args:
        if item not in kwargs['var_dict'].keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following REQUIRED Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    error_count = 0
    error_list = []
    for item in optional_args:
        if item not in kwargs['var_dict'].keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following Optional Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    # Load all required args values from kwargs
    error_count = 0
    error_list = []
    for item in kwargs['var_dict']:
        if item in required_args.keys():
            required_args[item] = kwargs['var_dict'][item]
            if required_args[item] == None:
                error_count =+ 1
                error_list += [item]

    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n - The Following REQUIRED Key(s) Argument(s) are Blank:\nPlease Validate "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    for item in kwargs['var_dict']:
        if item in optional_args.keys():
            optional_args[item] = kwargs['var_dict'][item]
    # Combine option and required dicts for Jinja template render
    polVars = {**required_args, **optional_args}
    return(polVars)

#======================================================
# Function - Read Excel Workbook Data
#======================================================
def read_in(excel_workbook, kwargs):
    try:
        kwargs['wb'] = load_workbook(excel_workbook)
        print("Workbook Loaded.")
    except Exception as e:
        print(f"Something went wrong while opening the workbook - {excel_workbook}... ABORT!")
        sys.exit(e)
    return kwargs

#======================================================
# Validate File is in Repo URL
#======================================================
def repo_url_test(file, pargs):
    repo_url = f'https://{pargs.repository_server}{pargs.repository_path}{file}'
    try:
        r = requests.head(repo_url, allow_redirects=True, verify=False, timeout=10)
    except requests.RequestException as e:
        print(f"!!! ERROR !!!\n  Exception when calling {repo_url}:\n {e}\n")
        sys.exit(1)
    return repo_url

#======================================================
# Function - Prompt User for Sensitive Values
#======================================================
def sensitive_var_value(kwargs):
    json_data = kwargs.json_data
    if kwargs.deployment_type:
        sensitive_var = kwargs.sensitive_var
    else: sensitive_var = 'TF_VAR_%s' % (kwargs.sensitive_var)
    # -------------------------------------------------------------------------------------------------------------------------
    # Check to see if the Variable is already set in the Environment, and if not prompt the user for Input.
    #--------------------------------------------------------------------------------------------------------------------------
    if os.environ.get(sensitive_var) is None:
        print(f"\n---------------------------------------------------------------------------------\n")
        print(f"  The Script did not find {sensitive_var} as an 'environment' variable.")
        print(f"  To not be prompted for the value of {kwargs.sensitive_var} each time")
        print(f"  add the following to your local environemnt:\n")
        print(f"    - Linux: export {sensitive_var}='{kwargs.sensitive_var}_value'")
        print(f"    - Windows: $env:{sensitive_var}='{kwargs.sensitive_var}_value'")
        print(f"\n---------------------------------------------------------------------------------\n")
    if os.environ.get(sensitive_var) is None and kwargs.sensitive_var == 'ipmi_key':
        print(f'\n---------------------------------------------------------------------------------\n')
        print(f'  The ipmi_key Must be in Hexidecimal Format [a-fA-F0-9]')
        print(f'  and no longer than 40 characters.')
        print(f'\n---------------------------------------------------------------------------------\n')
    if os.environ.get(sensitive_var) is None:
        valid = False
        while valid == False:
            varValue = input('press enter to continue: ')
            if varValue == '': valid = True
        valid = False
        while valid == False:
            if kwargs.get('Multi_Line_Input'):
                print(f"Enter the value for {kwargs.sensitive_var}:")
                lines = []
                while True:
                    # line = input('')
                    line = stdiomask.getpass(prompt='')
                    if line: lines.append(line)
                    else: break
                if not re.search('(certificate|private_key)', sensitive_var): secure_value = '\\n'.join(lines)
                else: secure_value = '\n'.join(lines)
            else:
                valid_pass = False
                while valid_pass == False:
                    password1 = stdiomask.getpass(prompt=f"Enter the value for {kwargs.sensitive_var}: ")
                    password2 = stdiomask.getpass(prompt=f"Re-Enter the value for {kwargs.sensitive_var}: ")
                    if password1 == password2:
                        secure_value = password1
                        valid_pass = True
                    else: print('!!! ERROR !!! Sensitive Values did not match.  Please re-enter...')

            # Validate Sensitive Passwords
            cert_regex = re.compile(r'^\-{5}BEGIN (CERTIFICATE|PRIVATE KEY)\-{5}.*\-{5}END (CERTIFICATE|PRIVATE KEY)\-{5}$')
            if re.search('(certificate|private_key)', sensitive_var):
                if not re.search(cert_regex, secure_value): valid = True
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print(f'    !!! ERROR !!! Invalid Value for the {sensitive_var}.  Please re-enter the {sensitive_var}.')
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            elif re.search('intersight_apikey', sensitive_var):
                minLength = 74
                maxLength = 74
                rePattern = '^[\\da-f]{24}/[\\da-f]{24}/[\\da-f]{24}$'
                varName = 'Intersight API Key'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'bind' in sensitive_var:
                jsonVars = json_data['iam.LdapBaseProperties'].allOf[1].properties
                minLength = 1
                maxLength = 254
                rePattern = jsonVars['Password']['pattern']
                varName = 'LDAP Binding Password'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'community' in sensitive_var:
                varName = 'SNMP Community String'
                valid = validating.snmp_string(varName, secure_value)
            elif 'ipmi_key' in sensitive_var: valid = validating.ipmi_key_check(secure_value)
            elif 'iscsi_boot' in sensitive_var:
                jsonVars = json_data['vnic.IscsiAuthProfile'].allOf[1].properties
                minLength= 12
                maxLength= 16
                rePattern= jsonVars.Password.pattern
                varName  = 'iSCSI Boot Password'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'local' in sensitive_var or 'ucs_password' in sensitive_var:
                jsonVars = json_data['iam.EndPointUserRole'].allOf[1].properties
                minLength= jsonVars.Password.minLength
                maxLength= jsonVars.Password.maxLength
                rePattern= jsonVars.Password.pattern
                varName  = 'Local User Password'
                if kwargs.get('enforce_strong_password'): enforce_pass = kwargs.enforce_strong_password
                else: enforce_pass = False
                if enforce_pass == True:
                    minLength = 8
                    maxLength = 20
                    valid = validating.strong_password(kwargs.sensitive_var, secure_value, minLength, maxLength)
                else: valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'secure_passphrase' in sensitive_var:
                jsonVars = json_data['memory.PersistentMemoryLocalSecurity'].allOf[1].properties
                minLength = jsonVars['SecurePassphrase']['minLength']
                maxLength = jsonVars['SecurePassphrase']['maxLength']
                rePattern = jsonVars['SecurePassphrase']['pattern']
                varName = 'Persistent Memory Secure Passphrase'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'snmp' in sensitive_var:
                print('matched snmp')
                exit()
                jsonVars = json_data['snmp.Policy'].allOf[1].properties
                minLength = 1
                maxLength = jsonVars['TrapCommunity']['maxLength']
                rePattern = '^[\\S]+$'
                if 'auth' in sensitive_var: varName = 'SNMP Authorization Password'
                else: varName = 'SNMP Privacy Password'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)
            elif 'vmedia' in sensitive_var:
                jsonVars = json_data['vmedia.Mapping'].allOf[1].properties
                minLength = 1
                maxLength = jsonVars['Password']['maxLength']
                rePattern = '^[\\S]+$'
                varName = 'vMedia Mapping Password'
                valid = validating.length_and_regex_sensitive(rePattern, varName, secure_value, minLength, maxLength)

        # Add Policy Variables to imm_dict
        if kwargs.get('org'):
            org = kwargs.org
            if not kwargs.imm_dict.orgs.get(org):
                kwargs.imm_dict.orgs[org] = DotMap()
                if not kwargs.imm_dict.orgs[org].get('sensitive_vars'):
                    kwargs.imm_dict.orgs[org].sensitive_vars = []
                kwargs.imm_dict.orgs[org].sensitive_vars.append(sensitive_var)

        # Add the Variable to the Environment
        os.environ[sensitive_var] = '%s' % (secure_value)
        kwargs.var_value = secure_value
    else:
        # Add the Variable to the Environment
        if kwargs.get('Multi_Line_Input'):
            var_value = os.environ.get(sensitive_var)
            kwargs.var_value = var_value.replace('\n', '\\n')
        else: kwargs.var_value = os.environ.get(sensitive_var)
    return kwargs

#======================================================
# Function - Wizard for SNMP Trap Servers
#======================================================
def snmp_trap_servers(kwargs):
    json_data   = kwargs['json_data']
    loop_count = 1
    kwargs['snmp_traps'] = []
    valid_traps = False
    while valid_traps == False:
        #==============================================
        # Get API Data
        #==============================================
        kwargs['multi_select'] = False
        jsonVars = json_data['snmp.Trap'].allOf[1].properties
        if len(kwargs['snmp_users']) == 0:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  There are no valid SNMP Users so Trap Destinations can only be set to SNMPv2.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            kwargs.sensitive_var = f'snmp_community_string_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
            #==============================================
            # Prompt User for SNMP Trap Type
            #==============================================
            kwargs['Description'] = jsonVars['Type']['description']
            kwargs['jsonVars'] = sorted(jsonVars['Type'].enum)
            kwargs['defaultVar'] = jsonVars['Type']['default']
            kwargs['varType'] = 'SNMP Trap Type'
            trap_type = variablesFromAPI(kwargs)
        else:
            #==============================================
            # Prompt User for SNMP Username
            #==============================================
            kwargs['Description'] = '    Please Select the SNMP User to assign to this Destination:\n'
            kwargs['var_type'] = 'SNMP User'
            snmp_users = []
            for item in kwargs['snmp_users']: snmp_users.append(item['name'])
            snmp_user = vars_from_list(snmp_users, kwargs)
            snmp_user = snmp_user[0]
        valid = False
        while valid == False:
            destination_address = input(f'What is the SNMP Trap Destination Hostname/Address? ')
            if not destination_address == '':
                if re.search(r'^[0-9a-fA-F]+[:]+[0-9a-fA-F]$', destination_address) or \
                    re.search(r'^(\d{1,3}\.){3}\d{1,3}$', destination_address):
                    valid = validating.ip_address('SNMP Trap Destination', destination_address)
                else: valid = validating.dns_name('SNMP Trap Destination', destination_address)
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error!! Invalid Value.  Please Re-enter the SNMP Trap Destination Hostname/Address.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
        valid = False
        while valid == False:
            port = input(f'Enter the Port to Assign to this Destination.  Valid Range is 1-65535.  [162]: ')
            if port == '': port = 162
            if re.search(r'[0-9]{1,4}', str(port)): valid = validating.snmp_port('SNMP Port', port, 1, 65535)
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Invalid Entry!  Please Enter a valid Port in the range of 1-65535.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
        if len(kwargs['snmp_users']) == 0:
            snmp_destination = {'community':loop_count, 'destination_address':destination_address, 'port':port}
            if not trap_type == 'Trap': snmp_destination.update({'trap_type':trap_type})
        else: snmp_destination = {'destination_address':destination_address, 'port':port, 'user':snmp_user}
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(textwrap.indent(yaml.dump(snmp_destination, Dumper=MyDumper, default_flow_style=False
        ), ' '*4, predicate=None))
        print(f'\n-------------------------------------------------------------------------------------------\n')
        valid_confirm = False
        while valid_confirm == False:
            confirm_v = input('Do you want to accept the above configuration?  Enter "Y" or "N" [Y]: ')
            if confirm_v == 'Y' or confirm_v == '':
                kwargs['snmp_traps'].append(snmp_destination)
                policy_type = 'SNMP Trap Destination'
                valid_exit = False
                while valid_exit == False:
                    loop_exit, valid_traps = exit_default(policy_type, 'N')
                    if loop_exit == True: loop_count += 1; valid_confirm = True; valid_exit = True
                    elif loop_exit == False: valid_confirm = True; valid_exit = True
            elif confirm_v == 'N':
                policy_type = 'SNMP Trap Server'
                message_starting_over(policy_type)
                valid_confirm = True
            else: message_invalid_y_or_n('long')
    return kwargs

#======================================================
# Function - Wizard for SNMP Users
#======================================================
def snmp_users(kwargs):
    loop_count = 1
    json_data = kwargs['json_data']
    kwargs['snmp_users'] = []
    valid_users = False
    while valid_users == False:
        kwargs['multi_select'] = False
        jsonVars = json_data['snmp.User'].allOf[1].properties
        snmpUser = False
        while snmpUser == False:
            #================================================
            # Prompt User for SNMP Username
            #================================================
            kwargs['jData'] = deepcopy(jsonVars['Name'])
            kwargs['jData']['default']  = 'snmpadmin'
            kwargs['jData']['pattern']  = '^([a-zA-Z]+[a-zA-Z0-9\\-\\_\\.\\@]+)$'
            kwargs['jData']['varInput'] = 'What is the SNMPv3 Username?'
            kwargs['jData']['varName']  = 'SNMP User'
            snmp_user = varStringLoop(kwargs)
            if snmp_user == 'admin':
                print(f'\n{"-"*91}\n')
                print(f'  Error!! Invalid Value.  admin may not be used for the snmp user value.')
                print(f'\n{"-"*91}\n')
            else: snmpUser = True
        #================================================
        # Prompt User for SNMP Security Level
        #================================================
        kwargs['jData'] = deepcopy(jsonVars['SecurityLevel'])
        kwargs['jData']['varType'] = 'SNMP Security Level'
        security_level = variablesFromAPI(kwargs)
        if security_level == 'AuthNoPriv' or security_level == 'AuthPriv':
            #================================================
            # Prompt User for SNMP Authentication Type
            #================================================
            kwargs['jData'] = deepcopy(jsonVars['AuthType'])
            kwargs['jData']['default'] = 'SHA'
            kwargs['jData']['popList'] = ['NA', 'SHA-224', 'SHA-256', 'SHA-384', 'SHA-512']
            kwargs['jData']['varType'] = 'SNMP Auth Type'
            auth_type = variablesFromAPI(kwargs)
        if security_level == 'AuthNoPriv' or security_level == 'AuthPriv':
            kwargs.sensitive_var = f'snmp_auth_password_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
        if security_level == 'AuthPriv':
            #================================================
            # Prompt User for SNMP Privacy Type
            #================================================
            kwargs['jData'] = deepcopy(jsonVars['PrivacyType'])
            kwargs['jData']['default'] = 'AES'
            kwargs['jData']['popList'] = ['NA']
            kwargs['jData']['varType'] = 'SNMP Auth Type'
            privacy_type = variablesFromAPI(kwargs)
            kwargs.sensitive_var = f'snmp_privacy_password_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
        snmp_user = {
            'auth_password':loop_count, 'auth_type':auth_type,
            'name':snmp_user, 'security_level':security_level
        }
        if security_level == 'AuthPriv':
            snmp_user.update({'privacy_password':loop_count, 'privacy_type':privacy_type})
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(textwrap.indent(yaml.dump(snmp_user, Dumper=MyDumper, default_flow_style=False
        ), " "*3, predicate=None))
        print(f'\n-------------------------------------------------------------------------------------------\n')
        valid_confirm = False
        while valid_confirm == False:
            confirm_v = input('Do you want to accept the above configuration?  Enter "Y" or "N" [Y]: ')
            if confirm_v == 'Y' or confirm_v == '':
                kwargs['snmp_users'].append(snmp_user)
                policy_type = 'SNMP User'
                #==============================================
                # Create Additional Policy or Exit Loop
                #==============================================
                valid_exit = False
                while valid_exit == False:
                    loop_exit, valid_users = exit_default(policy_type, 'N')
                    if loop_exit == True: loop_count += 1; valid_confirm = True; valid_exit = True
                    elif loop_exit == False: valid_confirm = True; valid_exit = True
            elif confirm_v == 'N':
                policy_type = 'SNMP User'
                message_starting_over(policy_type)
                valid_confirm = True
            else: message_invalid_y_or_n('long')
    return kwargs

#======================================================
# Function - Define stdout_log output
#======================================================
def stdout_log(ws, row_num):
    if log_level == 0: return
    elif ((log_level == (1) or log_level == (2)) and
            (ws) and (row_num is None)) and row_num == 'begin':
        print(f'-----------------------------------------------------------------------------\n')
        print(f'   Begin Worksheet "{ws.title}" evaluation...')
        print(f'\n-----------------------------------------------------------------------------\n')
    elif (log_level == (1) or log_level == (2)) and row_num == 'end':
        print(f'\n-----------------------------------------------------------------------------\n')
        print(f'   Completed Worksheet "{ws.title}" evaluation...')
        print(f'\n-----------------------------------------------------------------------------')
    elif log_level == (2) and (ws) and (row_num is not None):
        if re.fullmatch('[0-9]', str(row_num)):
            print(f'    - Evaluating Row   {row_num}...')
        elif re.fullmatch('[0-9][0-9]',  str(row_num)):
            print(f'    - Evaluating Row  {row_num}...')
        elif re.fullmatch('[0-9][0-9][0-9]',  str(row_num)):
            print(f'    - Evaluating Row {row_num}...')
    else: return

#======================================================
# Function - Wizard for Syslog Servers
#======================================================
def syslog_servers(kwargs):
    json_data = kwargs['json_data']
    kwargs['remote_logging'] = []
    policy_type = 'Syslog Server'
    syslog_count = 0
    syslog_loop = False
    while syslog_loop == False:
        jsonVars = json_data['syslog.RemoteClientBase'].allOf[1].properties
        #================================================
        # Prompt User for Syslog Server
        #================================================
        kwargs['jData'] = deepcopy(jsonVars['Hostname'])
        kwargs['jData']["pattern"]  = '.*'
        kwargs['jData']["varInput"] = 'What is the Hostname/IP of the Remote Server?'
        kwargs['jData']["varName"]  = 'Remote Logging Server'
        kwargs['jData']['varType']  = 'hostname'
        hostname = varStringLoop(kwargs)
        #================================================
        # Prompt User for Syslog Minimum Severity
        #================================================
        kwargs['jData'] = deepcopy(jsonVars['MinSeverity'])
        kwargs['jData']['varType'] = 'Minimum Severity To Report'
        min_severity = variablesFromAPI(kwargs)
        #================================================
        # Prompt User for Syslog Protocol
        #================================================
        kwargs['jData'] = deepcopy(jsonVars['Protocol'])
        kwargs['jData']['varType'] = 'Syslog Protocol'
        protocol = variablesFromAPI(kwargs)
        #================================================
        # Prompt User for LDAP Provider Port
        #================================================
        kwargs['jData'] = deepcopy(jsonVars['Port'])
        kwargs['jData']['varInput'] = f'What is Port for {hostname}?'
        kwargs['jData']['varName']  = 'LDAP Port'
        port = varNumberLoop(kwargs)
        remote_host = {'enable':True, 'hostname':hostname, 'minimum_severity':min_severity, 'port':port, 'protocol':protocol}
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(textwrap.indent(yaml.dump(remote_host, Dumper=MyDumper, default_flow_style=False
        ), " "*3, predicate=None))
        print(f'-------------------------------------------------------------------------------------------\n')
        valid_confirm = False
        while valid_confirm == False:
            confirm_host = input('Do you want to accept the configuration above?  Enter "Y" or "N" [Y]: ')
            if confirm_host == 'Y' or confirm_host == '':
                kwargs['remote_logging'].append(remote_host)
                if syslog_count == 1: syslog_loop = True; valid_confirm = True
                if syslog_count == 0:
                    valid_exit = False
                    while valid_exit == False:
                        loop_exit, syslog_loop = exit_default(policy_type, 'Y')
                        print(loop_exit, valid_confirm)
                        if loop_exit == False: syslog_count += 1; valid_confirm = True; valid_exit = True
                        elif loop_exit == True: valid_confirm = True; valid_exit = True
            elif confirm_host == 'N':
                message_starting_over(policy_type)
                valid_confirm = True
            else: message_invalid_y_or_n('long')
    return kwargs

#======================================================
# Function - Create a List of Subnet Hosts
#======================================================
def subnet_list(kwargs):
    ip_version = kwargs['ip_version']
    if ip_version == 'v4': prefix = kwargs['subnetMask']
    else: prefix = kwargs['prefix']
    gateway = kwargs['defaultGateway']
    subnetList = list(ipaddress.ip_network(f"{gateway}/{prefix}", strict=False).hosts())
    return subnetList

#======================================================
# Function - Format Terraform Files
#======================================================
def terraform_fmt(folder):
    # Run terraform fmt to cleanup the formating for all of the auto.tfvar files and tf files if needed
    print(f'\n-----------------------------------------------------------------------------\n')
    print(f'  Running "terraform fmt" in folder "{folder}",')
    print(f'  to correct variable formatting!')
    print(f'\n-----------------------------------------------------------------------------\n')
    p = subprocess.Popen(
        ['terraform', 'fmt', folder],
        stdout = subprocess.PIPE,
        stderr = subprocess.PIPE
    )
    print('Format updated for the following Files:')
    for line in iter(p.stdout.readline, b''):
        line = line.decode("utf-8")
        line = line.strip()
        print(f'- {line}')

#========================================================
# Function to Pull Latest Versions of Providers
#========================================================
def terraform_provider_config(kwargs):
    args     = kwargs['args']
    baseRepo = args.dir
    org      = kwargs['org']
    url_list = [
        'https://github.com/CiscoDevNet/terraform-provider-intersight/tags/',
        'https://github.com/hashicorp/terraform/tags',
        'https://github.com/netascode/terraform-provider-utils/tags/'
    ]
    for url in url_list:
        # Get the Latest Release Tag for the Provider
        r = requests.get(url, stream=True)
        repoVer = 'BLANK'
        stringMatch = False
        while stringMatch == False:
            for line in r.iter_lines():
                toString = line.decode("utf-8")
                if re.search(r'/releases/tag/v(\d+\.\d+\.\d+)\"', toString):
                    repoVer = re.search('/releases/tag/v(\d+\.\d+\.\d+)', toString).group(1)
                    break
            stringMatch = True
        
        # Make sure the latest_versions Key exists
        if kwargs.get('latest_versions') == None:
            kwargs['latest_versions'] = {}
        
        # Set Provider Version
        if   'intersight' in url:
            kwargs['latest_versions']['intersight_provider_version'] = repoVer
        elif 'netascode' in url:
            kwargs['latest_versions']['utils_provider_version'] = repoVer
        else: kwargs['latest_versions']['terraform_version'] = repoVer
    
    # Return kwargs
    return kwargs
    
#======================================================
# Function - Prompt User for Sensitive Variables
#======================================================
def tfc_sensitive_variables(varValue, json_data, polVars):
    polVars['Variable'] = varValue
    if 'ipmi_key' in varValue: polVars['Description'] = 'IPMI over LAN Encryption Key'
    elif 'iscsi' in varValue: polVars['Description'] = 'iSCSI Boot Password'
    elif 'local_user' in varValue: polVars['Description'] = 'Local User Password'
    elif 'access_comm' in varValue: polVars['Description'] = 'SNMP Access Community String'
    elif 'snmp_auth' in varValue: polVars['Description'] = 'SNMP Authorization Password'
    elif 'snmp_priv' in varValue: polVars['Description'] = 'SNMP Privacy Password'
    elif 'trap_comm' in varValue: polVars['Description'] = 'SNMP Trap Community String'
    polVars['varValue'] = sensitive_var_value(json_data, **polVars)
    polVars['varId'] = varValue
    polVars['varKey'] = varValue
    polVars['Sensitive'] = True
    print('* Adding "{}" to "{}"').format(polVars['Description'], polVars['workspaceName'])
    return polVars

#==========================================================
# Function - Prompt User for Chassis/Server Serial Numbers
#==========================================================
def ucs_serial(kwargs):
    baseRepo    = kwargs['args'].dir
    device_type = kwargs['device_type']
    org         = kwargs['org']
    yaml_file   = kwargs['yaml_file']
    valid = False
    while valid == False:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'  Note: If you do not have the Serial Number at this time you can manually add it to:')
        print(f'    - {baseRepo}{os.sep}{org}{os.sep}profiles{os.sep}{yaml_file}.yaml')
        print(f'      file later.')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        serial = input(f'What is the Serial Number of the {device_type}? [press enter to skip]: ')
        if serial == '': serial = 'unknown'; valid = True
        elif re.fullmatch(r'^[A-Z]{3}[2-3][\d]([0][1-9]|[1-4][0-9]|[5][1-3])[\dA-Z]{4}$', serial): valid = True
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  Error!! Invalid Serial Number.  "{serial}" is not a valid serial.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return serial

#======================================================
# Function - Prompt User for Domain Serial Numbers
#======================================================
def ucs_domain_serials(kwargs):
    baseRepo = kwargs['args'].dir
    org = kwargs['org']
    print(f'\n-------------------------------------------------------------------------------------------\n')
    print(f'  Note: If you do not have the Serial Numbers at this time you can manually add them here:\n')
    print(f'    * {baseRepo}{os.sep}{org}{os.sep}profiles{os.sep}domain.yaml\n')
    print(f'  After the Wizard has completed.')
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        polVars = {}
        fabrics = ['A','B']
        for x in fabrics:
            polVars[f'serial_{x}'] = input(f'What is the Serial Number of Fabric {x}? [press enter to skip]: ')
            if polVars[f'serial_{x}'] == '':
                polVars[f'serial_{x}'] = 'unknown'
                valid = True
            elif re.fullmatch(r'^[A-Z]{3}[2-3][\d]([0][1-9]|[1-4][0-9]|[5][1-3])[\dA-Z]{4}$', polVars[f'serial_{x}']):
                valid = True
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print('  Error!! Invalid Serial Number.  "{}" is not a valid serial.').format(polVars[f'serial_{x}'])
                print(f'\n-------------------------------------------------------------------------------------------\n')
    serials = [polVars['serial_A'], polVars['serial_B']]
    return serials

#========================================================
# Function to Validate Worksheet User Input
#========================================================
def validate_args(json_data, kwargs):
    json_data = kwargs['validateData']
    for i in json_data['required_args']:
        if json_data[i]['type'] == 'boolean':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                validating.boolean(i, kwargs)
        elif json_data[i]['type'] == 'hostname':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                if ':' in kwargs['var_dict'][i]:
                    validating.ip_address_ws(i, kwargs)
                elif re.search('[a-z]', kwargs['var_dict'][i], re.IGNORECASE):
                    validating.dns_name_ws(i, kwargs)
                else:
                    validating.ip_address_ws(i, kwargs)
        elif json_data[i]['type'] == 'list_of_email':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for email in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = email
                    validating.email_ws(f'{i}_{count}', kwargs)
        elif json_data[i]['type'] == 'email':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                validating.email_ws(i, kwargs)
        elif json_data[i]['type'] == 'integer':
            if kwargs['var_dict'][i] == None:
                kwargs['var_dict'][i] = json_data[i]['default']
            else:
                validating.number_check(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_domains':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for domain in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'domain_{count}'] = domain
                    validating.domain_ws(f'domain_{count}', kwargs)
                    kwargs['var_dict'].pop(f'domain_{count}')
                    count += 1
        elif json_data[i]['type'] == 'list_of_hosts':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for hostname in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = hostname
                    if ':' in hostname:
                        validating.ip_address_ws(f'{i}_{count}', kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE):
                        validating.dns_name_ws(f'{i}_{count}', kwargs)
                    else:
                        validating.ip_address_ws(f'{i}_{count}', kwargs)
                    kwargs['var_dict'].pop(f'{i}_{count}')
                    count += 1
        elif json_data[i]['type'] == 'list_of_integer':
            if kwargs['var_dict'][i] == None:
                kwargs['var_dict'][i] = json_data[i]['default']
            else:
                validating.number_list(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_string':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                validating.string_list(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_values':
            if kwargs['var_dict'][i] == None:
                kwargs['var_dict'][i] = json_data[i]['default']
            else:
                validating.list_values(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_vlans':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                validating.vlans(i, kwargs)
        elif json_data[i]['type'] == 'string':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                validating.string_pattern(i, json_data, kwargs)
        else:
            print(f"error validating.  Type not found {json_data[i]['type']}. 2.")
            exit()
    for i in json_data['optional_args']:
        if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
            if re.search(r'^module_[\d]+$', i):
                validating.list_values_key('modules', i, json_data, kwargs)
            elif json_data[i]['type'] == 'boolean':
                validating.boolean(i, json_data, kwargs)
            elif json_data[i]['type'] == 'domain':
                validating.domain_ws(i, kwargs)
            elif json_data[i]['type'] == 'list_of_email':
                count = 1
                for email in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = email
                    validating.email_ws(f'{i}_{count}', json_data, kwargs)
            elif json_data[i]['type'] == 'email':
                validating.email_ws(i, json_data, kwargs)
            elif json_data[i]['type'] == 'hostname':
                if ':' in kwargs['var_dict'][i]:
                    validating.ip_address_ws(i, kwargs)
                elif re.search('[a-z]', kwargs['var_dict'][i], re.IGNORECASE):
                    validating.dns_name_ws(i, kwargs)
                else:
                    validating.ip_address_ws(i, kwargs)
            elif json_data[i]['type'] == 'integer':
                validating.number_check(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_integer':
                validating.number_list(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_hosts':
                count = 1
                for hostname in kwargs['var_dict'][i].split(','):
                    kwargs[f'{i}_{count}'] = hostname
                    if ':' in hostname:
                        validating.ip_address_ws(f'{i}_{count}', kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE):
                        validating.dns_name_ws(f'{i}_{count}', kwargs)
                    else:
                        validating.ip_address_ws(f'{i}_{count}', kwargs)
                    kwargs['var_dict'].pop(f'{i}_{count}')
                    count += 1
            elif json_data[i]['type'] == 'list_of_macs':
                count = 1
                for mac in kwargs['var_dict'][i].split(','):
                    kwargs[f'{i}_{count}'] = mac
                    validating.mac_address(f'{i}_{count}', kwargs)
                    kwargs.pop(f'{i}_{count}')
                    count += 1
            elif json_data[i]['type'] == 'list_of_string':
                validating.string_list(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_values':
                validating.list_values(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_vlans':
                validating.vlans(i, kwargs)
            elif json_data[i]['type'] == 'mac_address':
                validating.mac_address(i, kwargs)
            elif json_data[i]['type'] == 'string':
                validating.string_pattern(i, json_data, kwargs)
            else:
                print(f"error validating.  Type not found {json_data[i]['type']}. 3.")
                exit()
    return kwargs

#======================================================
# Function - Check VLAN exists in VLAN Policy
#======================================================
def validate_vlan_in_policy(vlan_policy_list, vlan_id):
    valid = False
    while valid == False:
        vlan_count = 0
        for vlan in vlan_policy_list:
            if int(vlan_id) == 1:
                vlan_count = 1
                continue
            if int(vlan) == int(vlan_id):
                vlan_count = 1
                continue
        if vlan_count == 1: valid = True; return valid
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  VLAN {vlan_id} not found in the VLAN Policy List.  Please us a VLAN from the list below:')
            print(f'  {vlan_policy_list}')
            print(f'\n-------------------------------------------------------------------------------------------\n')
            return valid

#======================================================
# Function - Prompt User with List of Options
#======================================================
def variablesFromAPI(kwargs):
    if kwargs['jData'].get('default'): vDefault = kwargs['jData']['default']
    else: vDefault = ''
    if kwargs['jData'].get('dontsort'): jVars = kwargs['jData'].enum
    else: jVars = sorted(kwargs['jData'].enum)
    varDesc = kwargs['jData']['description']
    varType = kwargs['jData']['varType']
    valid = False
    while valid == False:
        json_vars = jVars
        if not kwargs['jData'].get('popList') == None:
            if len(kwargs['jData']['popList']) > 0:
                for x in kwargs['jData']['popList']:
                    for r in range(0, len(json_vars)):
                        if json_vars[r] == x:
                            json_vars.pop(r)
                            break
        print(f'\n-------------------------------------------------------------------------------------------\n')
        if '\n' in varDesc:
            varDesc = varDesc.split('\n')
            for line in varDesc:
                if '*' in line: print(textwrap.fill(f'{line}',width=88, subsequent_indent='    '))
                else: print(textwrap.fill(f'{line}',88))
        else: print(textwrap.fill(f'{varDesc}',88))
        print(f'\n    Select an Option Below:')
        for index, value in enumerate(json_vars):
            index += 1
            if value == vDefault: defaultIndex = index
            if index < 10: print(f'     {index}. {value}')
            else: print(f'    {index}. {value}')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        if kwargs['multi_select'] == True:
            if not vDefault == '':
                var_selection = input(f'Please Enter the Option Number(s) to Select for {varType}.  [{defaultIndex}]: ')
            else: var_selection = input(f'Please Enter the Option Number(s) to Select for {varType}: ')
        else:
            if not vDefault == '':
                var_selection = input(f'Please Enter the Option Number to Select for {varType}.  [{defaultIndex}]: ')
            else: var_selection = input(f'Please Enter the Option Number to Select for {varType}: ')
        if not vDefault == '' and var_selection == '': var_selection = defaultIndex
        if kwargs['multi_select'] == False and re.search(r'^[0-9]+$', str(var_selection)):
            for index, value in enumerate(json_vars):
                index += 1
                if int(var_selection) == index:
                    selection = value
                    valid = True
        elif kwargs['multi_select'] == True and re.search(r'(^[0-9]+$|^[0-9\-,]+[0-9]$)', str(var_selection)):
            var_list = vlan_list_full(var_selection)
            var_length = int(len(var_list))
            var_count = 0
            selection = []
            for index, value in enumerate(json_vars):
                index += 1
                for vars in var_list:
                    if int(vars) == index:
                        var_count += 1
                        selection.append(value)
            if var_count == var_length: valid = True
            else:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  The list of Vars {var_list} did not match the available list.')
                print(f'\n-------------------------------------------------------------------------------------------\n')
        else: message_invalid_selection()
    return selection

#======================================================
# Function - Prompt User for Boolean Question
#======================================================
def varBoolLoop(kwargs):
    varDesc  = kwargs['jData']['description']
    varInput = kwargs['jData']['varInput']
    varName  = kwargs['jData']['varName']
    if kwargs['jData']['default'] == True: varDefault = 'Y'
    else: varDefault = 'N'
    print(f'\n-------------------------------------------------------------------------------------------\n')
    if '\n' in varDesc:
        newDescr = varDesc.split('\n')
        for line in newDescr:
            if '*' in line: print(textwrap.fill(f'{line}',width=88, subsequent_indent='    '))
            else: print(textwrap.fill(f'{line}',88))
    else: print(textwrap.fill(f'{varDesc}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{varInput}  Enter "Y" or "N" [{varDefault}]: ')
        if varValue == '':
            if varDefault == 'Y': varValue = True
            elif varDefault == 'N': varValue = False
            valid = True
        elif varValue == 'N':
            varValue = False
            valid = True
        elif varValue == 'Y':
            varValue = True
            valid = True
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value of "{varValue}" is Invalid!!! Please enter "Y" or "N".')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function - Prompt User for Input Number
#======================================================
def varNumberLoop(kwargs):
    varDefault = kwargs['jData']['default']
    varDesc    = kwargs['jData']['description']
    maximum    = kwargs['jData']['maximum']
    minimum    = kwargs['jData']['minimum']
    varInput   = kwargs['jData']['varInput']
    varName    = kwargs['jData']['varName']
    if kwargs['jData'].get('varType'): varType = kwargs['jData']['varType']
    else: varType = 'undefined'

    print(f'\n-------------------------------------------------------------------------------------------\n')
    if '\n' in varDesc:
        newDescr = varDesc.split('\n')
        for line in newDescr:
            if '*' in line: print(textwrap.fill(f'{line}',width=88, subsequent_indent='    '))
            else: print(textwrap.fill(f'{line}',88))
    else: print(textwrap.fill(f'{varDesc}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = input(f'{varInput}  [{varDefault}]: ')
        if varValue == '': varValue = varDefault
        if re.fullmatch(r'^[0-9]+$', str(varValue)):
            if varType == 'SnmpPort':
                valid = validating.snmp_port(varName, varValue, minimum, maximum)
            else: valid = validating.number_in_range(varName, varValue, minimum, maximum)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value of "{varValue}" is Invalid!!! ')
            print(f'   Valid range is {minimum} to {maximum}.')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function - Prompt User for Sensitive Input String
#======================================================
def varSensitiveStringLoop(kwargs):
    jDict      = kwargs['jDict']
    varDescr   = jDict['description']
    varName    = kwargs['jData']['varName']
    varRegex   = kwargs['jData']['pattern']
    if kwargs['jData'].get('default'):  varDefault = kwargs['jData']['default']
    else: varDefault = ''
    if kwargs['jData'].get('maximum'):  maximum = kwargs['jData']['maximum']
    else: maximum = 0
    if kwargs['jData'].get('minimum'):  minimum = kwargs['jData']['minimum']
    else: minimum = 0
    print(f'\n-------------------------------------------------------------------------------------------\n')
    if '\n' in newDescr:
        newDescr = varDescr.split('\n')
        for line in newDescr:
            if '*' in line: print(textwrap.fill(f'{line}',width=88, subsequent_indent='    '))
            else: print(textwrap.fill(f'{line}',88))
    else: print(textwrap.fill(f"{kwargs['description']}",88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        varValue = stdiomask.getpass(f"{kwargs['varInput']} ")
        if not varValue == '':
            valid = validating.length_and_regex_sensitive(varRegex, varName, varValue, minimum, maximum)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value is Invalid!!! ')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function - Prompt User for Input String
#======================================================
def varStringLoop(kwargs):
    varDesc    = kwargs['jData']['description']
    varInput   = kwargs['jData']['varInput']
    varName    = kwargs['jData']['varName']
    varRegex   = kwargs['jData']['pattern']
    if kwargs['jData'].get('default'):  varDefault = kwargs['jData']['default']
    else: varDefault = ''
    if   kwargs['jData'].get('maximum'):   maximum = kwargs['jData']['maximum']
    elif kwargs['jData'].get('maxLength'): maximum = kwargs['jData']['maxLength']
    else: maximum = 0
    if   kwargs['jData'].get('minimum'):   minimum = kwargs['jData']['minimum']
    elif kwargs['jData'].get('minLength'): minimum = kwargs['jData']['minLength']
    else: minimum = 0
    if kwargs['jData'].get('varType'): varType = kwargs['jData']['varType']
    else: varType = 'undefined'
    print(f'\n-------------------------------------------------------------------------------------------\n')
    if '\n' in varDesc:
        newDescr = varDesc.split('\n')
        for line in newDescr:
            if '*' in line: print(textwrap.fill(f'{line}',width=88, subsequent_indent='    '))
            else: print(textwrap.fill(f'{line}',88))
    else: print(textwrap.fill(f'{varDesc}',88))
    print(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        if not varDefault == '': varValue = input(f'{varInput}  [{varDefault}]: ')
        else: varValue = input(f'{varInput} ')
        if 'press enter to skip' in varInput and varValue == '': valid = True
        elif not varDefault == '' and varValue == '':
            varValue = varDefault
            valid = True
        elif not varValue == '':
            if varType == 'hostname':
                if re.search(r'^[a-zA-Z0-9]:', varValue):
                    valid = validating.ip_address(varName, varValue)
                if re.search(r'[a-zA-Z]', varValue):
                    valid = validating.dns_name(varName, varValue)
                elif re.search(r'^([0-9]{1,3}\.){3}[0-9]{1,3}$', varValue):
                    valid = validating.ip_address(varName, varValue)
                else:
                    print(f'\n-------------------------------------------------------------------------------------------\n')
                    print('  "{}" is not a valid address.').format(varValue)
                    print(f'\n-------------------------------------------------------------------------------------------\n')
            elif varType == 'list':
                varValue = varValue.split(',')
                for i in varValue:
                    valid_item = validating.length_and_regex(varRegex, varName, i, minimum, maximum)
                    if valid_item == False: valid = False; break
                if valid_item == True:
                    valid = True
            elif varType == 'url':
                if re.search('^http', varValue): valid = validating.url(varName, varValue)
                else:
                    varUrl = f'http://{varValue}'
                    valid = validating.url(varName, varUrl)
            else: valid = validating.length_and_regex(varRegex, varName, varValue, minimum, maximum)
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'   {varName} value of "{varValue}" is Invalid!!! ')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return varValue

#======================================================
# Function - Prompt User with Names for Policies
#======================================================
def vars_from_list(var_options, kwargs):
    selection = []
    selection_count = 0
    valid = False
    while valid == False:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f"{kwargs['Description']}")
        for index, value in enumerate(var_options):
            index += 1
            if index < 10: print(f'     {index}. {value}')
            else: print(f'    {index}. {value}')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        exit_answer = False
        while exit_answer == False:
            var_selection = input(f"Please Enter the Option Number to Select for {kwargs['var_type']}: ")
            if not var_selection == '':
                if re.search(r'[0-9]+', str(var_selection)):
                    xcount = 1
                    for index, value in enumerate(var_options):
                        index += 1
                        if int(var_selection) == index:
                            selection.append(value)
                            xcount = 0
                    if xcount == 0:
                        if selection_count % 2 == 0 and kwargs['multi_select'] == True:
                            answer_finished = input('Would you like to add another port to the'\
                                f" {kwargs['port_type']}?"'  Enter "Y" or "N" [Y]: ')
                        elif kwargs['multi_select'] == True:
                            answer_finished = input('Would you like to add another port to the'\
                                f" {kwargs['port_type']}?"'  Enter "Y" or "N" [N]: ')
                        elif kwargs['multi_select'] == False:
                            answer_finished = 'N'
                        if (selection_count % 2 == 0 and answer_finished == '') or answer_finished == 'Y':
                            exit_answer = True
                            selection_count += 1
                        elif answer_finished == '' or answer_finished == 'N':
                            exit_answer = True
                            valid = True
                        elif kwargs['multi_select'] == False:
                            exit_answer = True
                            valid = True
                        else: message_invalid_y_or_n('short')
                    else: message_invalid_selection()
                else: message_invalid_selection()
            else: message_invalid_selection()
    return selection

#======================================================
# Function - Collapse VLAN List
#======================================================
def vlan_list_format(vlan_list_expanded):
    vlan_list  = sorted(vlan_list_expanded)
    vlanGroups = itertools.groupby(vlan_list, key=lambda item, c=itertools.count():item-next(c))
    tempvlans  = [list(g) for k, g in vlanGroups]
    vlanList   = [str(x[0]) if len(x) == 1 else "{}-{}".format(x[0],x[-1]) for x in tempvlans]
    vlan_list  = ",".join(vlanList)
    return vlan_list

#======================================================
# Function - Expand VLAN List
#======================================================
def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
                a,b = v.split('-')
                a = int(a)
                b = int(b)
                vrange = range(a,b+1)
                for vl in vrange: full_vlan_list.append(int(vl))
            elif re.fullmatch('^\\d{1,4}$', v): full_vlan_list.append(int(v))
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-')
        a = int(a)
        b = int(b)
        vrange = range(a,b+1)
        for v in vrange: full_vlan_list.append(int(v))
    else: full_vlan_list.append(vlan_list)
    return full_vlan_list

#======================================================
# Function - To Request Native VLAN
#======================================================
def vlan_native_function(vlan_policy_list, vlan_list):
    native_count = 0
    nativeVlan = ''
    nativeValid = False
    while nativeValid == False:
        nativeVlan = input('Do you want to Configure one of the VLANs as a Native VLAN?  [press enter to skip]:')
        if nativeVlan == '': nativeValid = True
        else:
            for vlan in vlan_policy_list:
                if int(nativeVlan) == int(vlan):
                    native_count = 1
                    break
            if not native_count == 1: message_invalid_native_vlan(nativeVlan, vlan_list)
            else: nativeValid = True
    return nativeVlan

#======================================================
# Function - Prompt for VLANs and Configure Policy
#======================================================
def vlan_pool(name):
    valid = False
    while valid == False:
        print(f'\n-------------------------------------------------------------------------------------------\n')
        print(f'  The allowed vlan list can be in the format of:')
        print(f'     5 - Single VLAN')
        print(f'     1-10 - Range of VLANs')
        print(f'     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
        print(f'     1-10,20-30 - Ranges and Lists of VLANs')
        print(f'\n-------------------------------------------------------------------------------------------\n')
        VlanList = input(f'Enter the VLAN or List of VLANs to assign to the Domain VLAN Pool {name}: ')
        if not VlanList == '':
            vlanListExpanded = vlan_list_full(VlanList)
            valid_vlan = True
            for vlan in vlanListExpanded:
                valid_vlan = validating.number_in_range('VLAN ID', vlan, 1, 4094)
                if valid_vlan == False:
                    break
            if valid_vlan == False:
                print(f'\n-------------------------------------------------------------------------------------------\n')
                print(f'  Error with VLAN(s) assignment!!! VLAN List: "{VlanList}" is not Valid.')
                print(f'  The allowed vlan list can be in the format of:')
                print(f'     5 - Single VLAN')
                print(f'     1-10 - Range of VLANs')
                print(f'     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
                print(f'     1-10,20-30 - Ranges and Lists of VLANs')
                print(f'\n-------------------------------------------------------------------------------------------\n')
            else: valid = True
        else:
            print(f'\n-------------------------------------------------------------------------------------------\n')
            print(f'  The allowed vlan list can be in the format of:')
            print(f'     5 - Single VLAN')
            print(f'     1-10 - Range of VLANs')
            print(f'     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
            print(f'     1-10,20-30 - Ranges and Lists of VLANs')
            print(f'\n-------------------------------------------------------------------------------------------\n')
    return VlanList,vlanListExpanded

#========================================================
# Function to Determine which sites to write files to.
#========================================================
def write_to_repo_folder(polVars, kwargs):
    baseRepo   = kwargs['args'].dir
    dest_file  = kwargs['dest_file']
    # Setup jinja2 Environment
    template_path = pkg_resources.resource_filename(f'policies', 'templates/')
    templateLoader = jinja2.FileSystemLoader(searchpath=(template_path + 'provider/'))
    templateEnv = jinja2.Environment(loader=templateLoader)
    # Define the Template Source
    template = templateEnv.get_template(kwargs['template_file'])
    # Make sure the Destination Path and Folder Exist
    if not os.path.isdir(os.path.join(baseRepo)):
        dest_path = f'{os.path.join(baseRepo)}'
        os.makedirs(dest_path)
    dest_dir = os.path.join(baseRepo)
    if not os.path.exists(os.path.join(dest_dir, dest_file)):
        create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'
        os.system(create_file)
    tf_file = os.path.join(dest_dir, dest_file)
    wr_file = open(tf_file, 'w')

    # Render Payload and Write to File
    polVars = json.loads(json.dumps(polVars))
    polVars = {'keys':polVars}
    payload = template.render(polVars)
    wr_file.write(payload)
    wr_file.close()
