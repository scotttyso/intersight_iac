#=============================================================================
# Source Modules
#=============================================================================
def prRed(skk): print("\033[91m {}\033[00m" .format(skk))
import sys
try:
    from classes import pcolor, validating
    from copy import deepcopy
    from datetime import datetime, timedelta
    from dotmap import DotMap
    from git import cmd, Repo
    from openpyxl import load_workbook
    import ipaddress, itertools, jinja2, json, os, pexpect, pkg_resources, pytz, re, requests
    import shutil, subprocess, stdiomask, string, textwrap, validators, yaml
except ImportError as e:
    prRed(f'!!! ERROR !!!\n{e.__class__.__name__}')
    prRed(f" Module {e.name} is required to run this script")
    prRed(f" Install the module using the following: `pip install {e.name}`")
    sys.exit(1)

# Log levels 0 = None, 1 = Class only, 2 = Line
log_level = 2

# Exception Classes
class InsufficientArgs(Exception): pass
class MyDumper(yaml.Dumper):
    def increase_indent(self, flow=False, indentless=False):
        return super(MyDumper, self).increase_indent(flow, False)

#=====================================================
# pexpect - Login Function
#=====================================================
def child_login(kwargs):
    kwargs.sensitive_var = kwargs.password
    kwargs   = sensitive_var_value(kwargs)
    password = kwargs.var_value
    #=====================================================
    # Use 
    #=====================================================
    if kwargs.op_system == 'Windows':
        from pexpect import popen_spawn
        child = popen_spawn.PopenSpawn('cmd', encoding='utf-8', timeout=1)
    else:
        system_shell = os.environ['SHELL']
        child = pexpect.spawn(system_shell, encoding='utf-8')
    child.logfile_read = sys.stdout
    if kwargs.op_system == 'Windows':
        child.sendline(f'ping -n 2 {kwargs.hostname}')
        child.expect(f'ping -n 2 {kwargs.hostname}')
        child.expect_exact("> ")
        child.sendline(f'ssh {kwargs.username}@{kwargs.hostname} | Tee-Object Logs\{kwargs.hostname}.txt')
        child.expect(f'Tee-Object Logs\{kwargs.hostname}.txt')
    else:
        child.sendline(f'ping -c 2 {kwargs.hostname}')
        child.expect(f'ping -c 2 {kwargs.hostname}')
        child.expect_exact("$ ")
        child.sendline(f'ssh {kwargs.username}@{kwargs.hostname} | tee Logs/{kwargs.hostname}.txt')
        child.expect(f'tee Logs/{kwargs.hostname}.txt')
    logged_in = False
    while logged_in == False:
        i = child.expect(
            ['Are you sure you want to continue', 'closed', 'password:', 'Password:', kwargs.host_prompt, pexpect.TIMEOUT])
        if i == 0: child.sendline('yes')
        elif i == 1:
            prRed(f'\n!!! FAILED !!! to connect.  '\
                f'Please Validate {kwargs.hostname} is correct and username {kwargs.username} is correct.')
            sys.exit(1)
        elif i == 2: child.sendline(password)
        elif i == 3: child.sendline(password)
        elif i == 4: logged_in = True
        elif i == 5:
            prRed(f"\n{'-'*91}\n")
            prRed(f'!!! FAILED !!!\n Could not open SSH Connection to {kwargs.hostname}')
            prRed(f"\n{'-'*91}\n")
            sys.exit(1)
    return child, kwargs

#======================================================
# Function - Format Policy Description
#======================================================
def choose_policy(policy_type, kwargs):
    policy_descr = mod_pol_description(policy_type)
    policy_list = []
    for i in kwargs['policies'][policy_type]: policy_list.append(i['name'])
    valid = False
    while valid == False:
        pcolor.Cyan(f'\n{"-"*91}\n')
        if kwargs.get('optional_message'): pcolor.Cyan(kwargs['optional_message'])
        pcolor.Cyan(f'  {policy_descr} Policy Options:')
        for i, v in enumerate(policy_list):
            i += 1
            if i < 10: pcolor.Cyan(f'     {i}. {v}')
            else: pcolor.Cyan(f'    {i}. {v}')
        if kwargs['allow_opt_out'] == True: pcolor.Cyan(f'     99. Do not assign a(n) {policy_descr}.')
        pcolor.Cyan(f'     100. Create a New {policy_descr}.')
        pcolor.Cyan(f'\n{"-"*91}\n')
        policyOption = input(f"Select the Option Number for the {policy_descr} Policy to Assign to {kwargs['name']} Policy: ")
        if re.search(r'^[0-9]{1,3}$', policyOption):
            for i, v in enumerate(policy_list):
                i += 1
                if   int(policyOption) == i:   kwargs['policy'] = v;  valid = True; return kwargs
                elif int(policyOption) == 99:  kwargs['policy'] = ''; valid = True; return kwargs
                elif int(policyOption) == 100: kwargs['policy'] = 'create_policy'; valid = True; return kwargs
            if   int(policyOption) == 99:  kwargs['policy'] = ''; valid = True; return kwargs
            elif int(policyOption) == 100: kwargs['policy'] = 'create_policy'; valid = True; return kwargs
        else: message_invalid_selection()

#======================================================
# Function - Count the Number of Keys
#======================================================
def countKeys(ws, func):
    count = 0
    for i in ws.rows:
        if any(i):
            if str(i[0].value) == func: count += 1
    return count

#==========================================================
# Function for Processing easyDict and Creating YAML Files
#==========================================================
def create_yaml(orgs, kwargs):
    ezdata  = kwargs.ezdata.ezimm_class.properties
    classes = kwargs.ezdata.ezimm_class.properties.classes.enum
    def write_file(dest_dir, dest_file, dict, title1):
        if not os.path.exists(os.path.join(dest_dir, dest_file)):
            create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'
            os.system(create_file)
        wr_file = open(os.path.join(dest_dir, dest_file), 'w')
        wr_file.write('---\n')
        wr_file = open(os.path.join(dest_dir, dest_file), 'a')
        dash_length = '='*(len(title1) + 20)
        wr_file.write(f'#{dash_length}\n')
        wr_file.write(f'#   {title1} - Variables\n')
        wr_file.write(f'#{dash_length}\n')
        wr_file.write(yaml.dump(dict, Dumper=MyDumper, default_flow_style=False))
        wr_file.close()
    for item in classes:
        dest_dir = os.path.join(kwargs.args.dir, ezdata[item].directory)
        if item == 'policies':
            if not os.path.isdir(dest_dir): os.makedirs(dest_dir)
            for i in ezdata[item].enum:
                idict = DotMap()
                for org in orgs:
                    if not idict.get(org): idict[org] = DotMap()
                    for x in ezdata[i].enum:
                        if kwargs.imm_dict.orgs[org].get(item):
                            if kwargs.imm_dict.orgs[org][item].get(x):
                                idict[org][item][x] = deepcopy(kwargs.imm_dict.orgs[org][item][x])
                            if not len(idict[org][item][x]) > 0: idict[org][item].pop(x)
                    if len(idict[org][item]) > 0:
                        idict = json.dumps(idict.toDict())
                        idict = json.loads(idict)
                        for x in ezdata[i].enum:
                            if kwargs.imm_dict.orgs[org][item].get(x):
                                if type(idict[org][item][x]) == list:
                                    if idict[org][item][x][0].get('name'):
                                        idict[org][item][x] = list({v['name']:v for v in idict[org][item][x]}.values())
                                    elif idict[org][item][x][0].get('names'):
                                        idict[org][item][x] = list({v['names'][0]:v for v in idict[org][item][x]}.values())
                        if re.search('policies|pools|profiles|templates', dest_dir): dest_file = f"{i}.ezi.yaml"
                        else: dest_file = f"{i}.yaml"
                        title1 = f"{str.title(item)} -> {i}"
                        write_file(dest_dir, dest_file, idict, title1)
        else:
            if not os.path.isdir(dest_dir): os.makedirs(dest_dir)
            for i in ezdata[item].enum:
                idict = deepcopy(DotMap())
                if item == i:
                    for org in orgs:
                        if kwargs.imm_dict.orgs[org].get(item):
                            if len(kwargs.imm_dict.orgs[org][item]) > 0:
                                itemDict = deepcopy(kwargs.imm_dict.orgs[org][item].toDict())
                                idict[org][item] = itemDict
                                idict = json.dumps(idict.toDict())
                                idict = json.loads(idict)
                                if type(idict[org][item]) == list: idict[org][item] = list({v['name']:v for v in value}.values())
                                else:
                                    newdict = deepcopy(idict)
                                    if re.search('(netapp|storage)', item):
                                        for key, value in newdict[org][item].items():
                                            idict[org][item][key] = list({v['name']:v for v in value}.values())
                                    else:
                                        for key, value in newdict[org][item].items():
                                            if not 'name' in value: idict[org][item][key] = value
                                            else: idict[org][item][key] = list({v['name']:v for v in value}.values())
                                if re.search('policies|pools|profiles|templates', dest_dir): dest_file = f"{i}.ezi.yaml"
                                else: dest_file = f"{i}.yaml"
                                title1 = str.title(item.replace('_', ' '))
                                write_file(dest_dir, dest_file, idict, title1)
                else:
                    for org in orgs:
                        if kwargs.imm_dict.orgs[org].get(item):
                            if kwargs.imm_dict.orgs[org][item].get(i):
                                if len(kwargs.imm_dict.orgs[org][item][i]) > 0:
                                    idict[org][item][i] = deepcopy(kwargs.imm_dict.orgs[org][item][i])
                                    idict = json.dumps(idict.toDict())
                                    idict = json.loads(idict)
                                    if type(idict[org][item][i]) == list:
                                        if re.search('(chassis|server)', i) and item == 'profiles': idict[org][item][i] = list({
                                                v['targets'][0]['name']:v for v in idict[org][item][i]}.values())
                                        else: idict[org][item][i] = list({v['name']:v for v in idict[org][item][i]}.values())
                                    else:
                                        for a, b in idict[org][item][i].items(): b = list({v['name']:v for v in b}.values())
                                    if re.search('policies|pools|profiles|templates', dest_dir): dest_file = f"{i}.ezi.yaml"
                                    else: dest_file = f"{i}.yaml"
                                    title1 = f"{str.title(item.replace('_', ' '))} -> {str.title(i.replace('_', ' '))}"
                                    write_file(dest_dir, dest_file, idict, title1)
                        
#==========================================================
# Function for Processing Wizard Data and Creating YAML Files
#==========================================================
def create_wizard_yaml(kwargs):
    org = kwargs.org
    def write_file(dest_dir, dest_file, dict, title1):
        if not os.path.exists(os.path.join(dest_dir, dest_file)):
            create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'
            os.system(create_file)
        wr_file = open(os.path.join(dest_dir, dest_file), 'w')
        wr_file.write('---\n')
        wr_file = open(os.path.join(dest_dir, dest_file), 'a')
        dash_length = '='*(len(title1) + 20)
        wr_file.write(f'#{dash_length}\n')
        wr_file.write(f'#   {title1} - Variables\n')
        wr_file.write(f'#{dash_length}\n')
        wr_file.write(yaml.dump(dict, Dumper=MyDumper, default_flow_style=False))
        wr_file.close()
    dest_dir = os.path.join(kwargs.args.dir, 'wizard')
    if not os.path.isdir(dest_dir): os.makedirs(dest_dir)
    idict = deepcopy(DotMap())
    item = 'wizard'
    if kwargs.imm_dict.orgs[org].get(item):
        if len(kwargs.imm_dict.orgs[org][item]) > 0:
            itemDict = deepcopy(kwargs.imm_dict.orgs[org][item].toDict())
            idict[org][item] = itemDict
            idict = json.dumps(idict.toDict())
            idict = json.loads(idict)
            if type(idict[org][item]) == list: idict[org][item] = list({v['name']:v for v in value}.values())
            else:
                newdict = deepcopy(idict)
                for key, value in newdict[org][item].items():
                    if type(value) == str: idict[org][item][key] = value
                    else: idict[org][item][key] = list({v['name']:v for v in value}.values())
            dest_file = f"{org}_{item}.yaml"
            title1 = str.title(item.replace('_', ' '))
            write_file(dest_dir, dest_file, idict, title1)
                        
#=====================================================
# Determine if Timezone Uses Daylight Savings
#=====================================================
def disable_daylight_savings(zonename):
    tz = pytz.timezone(zonename)
    june = pytz.utc.localize(datetime(2023, 6, 2, 12, 1, tzinfo=None))
    december = pytz.utc.localize(datetime(2023, 12, 2, 12, 1, tzinfo=None))
    june_dst = june.astimezone(tz).dst() != timedelta(0)
    dec_dst  = december.astimezone(tz).dst() != timedelta(0)
    if june_dst == True and dec_dst == False: return False
    elif june_dst == False and dec_dst == True: return False
    elif june_dst == False and dec_dst == False: return True
    else:
        print(f'unknown Timezone Result for {zonename}')
        sys.exit(1)

#======================================================
# Function - Prompt User with question
#======================================================
def exit_confirm_loop(kwargs):
    question = 'Y'
    valid_confirm = False
    while valid_confirm == False:
        question = input(f'Do you want to accept the above configuration?  Enter "Y" or "N" [{kwargs.yes_or_no}]: ')
        if question == '': question = kwargs.yes_or_no
        if question == 'Y':
            kwargs.accept_configuration = True
            valid_exit = False
            while valid_exit == False:
                loop_exit = input(
                    f'Would You like to Configure another {kwargs.policy_type}?  Enter "Y" or "N" [{kwargs.yes_or_no}]: ')
                if loop_exit == '': loop_exit = 'N'
                if loop_exit == 'Y':   valid_confirm = True; valid_exit = True; kwargs.configure_additional = True
                elif loop_exit == 'N': valid_confirm = True; valid_exit = True; kwargs.configure_additional = False
                else: message_invalid_y_or_n('short')
        elif question == 'N':
            kwargs.accept_configuration = False
            kwargs.configure_additional = True
            message_starting_over(kwargs.policy_type)
            valid_confirm = True
        else: message_invalid_y_or_n('long')
    return kwargs

#======================================================
# Function - Ask User to Configure Additional Policy
#======================================================
def exit_default(policy_type, y_or_n):
    valid_exit = False
    while valid_exit == False:
        exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [{y_or_n}]: ')
        if exit_answer == '':    exit_answer = y_or_n
        if exit_answer == 'N':   configure_loop = True;  policy_loop = True;  valid_exit = True
        elif exit_answer == 'Y': configure_loop = False; policy_loop = False; valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, policy_loop

#======================================================
# Function - Ask User to Configure Additional Policy
#======================================================
def exit_default_del_tfc(policy_type, y_or_n):
    valid_exit = False
    while valid_exit == False:
        exit_answer = input(f'Would You like to {policy_type}?  Enter "Y" or "N" [{y_or_n}]: ')
        if exit_answer == '':    exit_answer = y_or_n
        if exit_answer == 'N':   policy_loop = True;  configure_loop = True;  valid_exit = True
        elif exit_answer == 'Y': policy_loop = False; configure_loop = False; valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, policy_loop

#======================================================
# Function - Prompt User with question
#======================================================
def exit_loop_default_yes(loop_count, policy_type):
    valid_exit = False
    while valid_exit == False:
        if loop_count % 2 == 0:
            exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [Y]: ')
        else: exit_answer = input(f'Would You like to Configure another {policy_type}?  Enter "Y" or "N" [N]: ')
        if (loop_count % 2 == 0 and exit_answer == '') or exit_answer == 'Y':
            configure_loop = False; loop_count += 1; policy_loop = False; valid_exit = True
        elif not loop_count % 2 == 0 and exit_answer == '':
            configure_loop = True;  loop_count += 1; policy_loop = True;  valid_exit = True
        elif exit_answer == 'N':
            configure_loop = True;  loop_count += 1; policy_loop = True;  valid_exit = True
        else: message_invalid_y_or_n('short')
    return configure_loop, loop_count, policy_loop

#========================================================
# Function to Append the imm_dict Dictionary
#========================================================
def ez_append(polVars, kwargs):
    class_path= kwargs['class_path']
    p         = class_path.split(',')
    org       = kwargs['org']
    polVars   = ez_remove_empty(polVars)
    polVars   = DotMap(polVars)
    # Confirm the Key Exists
    if not kwargs.imm_dict.orgs.get(org):
        kwargs.imm_dict.orgs[org] = DotMap()
    if len(p) >= 2:
        if not kwargs.imm_dict.orgs[org].get(p[0]):
            kwargs.imm_dict.orgs[org][p[0]] = DotMap()
    if len(p) >= 3:
        if not kwargs.imm_dict.orgs[org][p[0]].get(p[1]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]] = DotMap()
    if len(p) >= 4:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]].get(p[2]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]] = DotMap()
    if len(p) == 2:
        if not kwargs.imm_dict.orgs[org][p[0]].get(p[1]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]] = [deepcopy(polVars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]].append(deepcopy(polVars))
    elif len(p) == 3:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]].get(p[2]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]] = [deepcopy(polVars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].append(deepcopy(polVars))
    elif len(p) == 4:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].get(p[3]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]] = [deepcopy(polVars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]].append(deepcopy(polVars))
    elif len(p) == 5:
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]].get(p[3]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]] = DotMap()
        if not kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]].get(p[4]):
            kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]][p[4]] = [deepcopy(polVars)]
        else: kwargs.imm_dict.orgs[org][p[0]][p[1]][p[2]][p[3]][p[4]].append(deepcopy(polVars))

    return kwargs

#========================================================
# Function to Append the imm_dict Dictionary
#========================================================
def ez_append_wizard(polVars, kwargs):
    class_path = kwargs['class_path']
    p = class_path.split(',')
    polVars = ez_remove_empty(polVars)
    # Confirm the Key Exists
    if not kwargs.imm_dict.get('wizard'): kwargs.imm_dict['wizard'] = {}
    if len(p) >= 2:
        if not kwargs.imm_dict['wizard'].get(p[0]):
            kwargs.imm_dict['wizard'].update(deepcopy({p[0]:{}}))
    if len(p) >= 3:
        if not kwargs.imm_dict['wizard'][p[0]].get(p[1]):
            kwargs.imm_dict['wizard'][p[0]].update(deepcopy({p[1]:{}}))
    if len(p) == 1:
        if not kwargs.imm_dict['wizard'].get(p[0]):
            kwargs.imm_dict['wizard'].update(deepcopy({p[0]:[]}))
    elif len(p) == 2:
        if not kwargs.imm_dict['wizard'][p[0]].get(p[1]):
            kwargs.imm_dict['wizard'][p[0]].update(deepcopy({p[1]:[]}))
    elif len(p) == 3:
        if not kwargs.imm_dict['wizard'][p[0]][p[1]].get(p[2]):
            kwargs.imm_dict['wizard'][p[0]][p[1]].update(deepcopy({p[2]:[]}))
    # append the Dictionary
    if len(p) == 1: kwargs.imm_dict['wizard'][p[0]].append(deepcopy(polVars))
    if len(p) == 2: kwargs.imm_dict['wizard'][p[0]][p[1]].append(deepcopy(polVars))
    elif len(p) == 3: kwargs.imm_dict['wizard'][p[0]][p[1]][p[2]].append(deepcopy(polVars))
    return kwargs

#========================================================
# Function to Remove Empty Arguments
#========================================================
def ez_remove_empty(polVars):
    pop_list = []
    for k,v in polVars.items():
        if v == None: pop_list.append(k)
    for i in pop_list: polVars.pop(i)
    return polVars

#======================================================
# Function - find the Keys for each Section
#======================================================
def findKeys(ws, func_regex):
    func_list = {}
    for i in ws.rows:
        if any(i):
            if re.search(func_regex, str(i[0].value)): func_list.add(str(i[0].value))
    func_list = DotMap(dict(sorted(func_list.items())))
    return func_list

#======================================================
# Function - Assign the Variables to the Keys
#======================================================
def findVars(ws, func, rows, count):
    var_list = []
    var_dict = {}
    for i in range(1, rows + 1):
        if (ws.cell(row=i, column=1)).value == func:
            try:
                for x in range(2, 34):
                    if (ws.cell(row=i - 1, column=x)).value: var_list.append(str(ws.cell(row=i - 1, column=x).value))
                    else: x += 1
            except Exception as e: e = e; pass
            break
    vcount = 1
    while vcount <= count:
        var_dict[vcount] = {}
        var_count = 0
        for z in var_list: var_dict[vcount][z] = ws.cell(row=i + vcount - 1, column=2 + var_count).value; var_count += 1
        var_dict[vcount]['row'] = i + vcount - 1
        vcount += 1
    return var_dict

#========================================================
# Function - Prompt User for the Intersight Configurtion
#========================================================
def intersight_config(kwargs):
    kwargs.jData = DotMap()
    if kwargs.args.intersight_api_key_id == None:
        kwargs.sensitive_var = 'intersight_api_key_id'
        kwargs = sensitive_var_value(kwargs)
        kwargs.args.intersight_api_key_id = kwargs.var_value

    #==============================================
    # Prompt User for Intersight SecretKey File
    #==============================================
    secret_path = kwargs.args.intersight_secret_key
    if not re.search('BEGIN RSA PRIVATE KEY.*END RSA PRIVATE KEY', secret_path):
        secret_loop = False
        while secret_loop == False:
            valid = False
            if secret_path == None:
                varName = 'intersight_secret_key'
                pcolor.Cyan(f'\n{"-"*91}\n\n  The Script did not find {varName} as an `environment` variable.')
                pcolor.Cyan(f'  To not be prompted for the value of {varName} each time\n  add the following to your local environemnt:')
                pcolor.Cyan(f"    - Linux: export {varName}='{varName}_value'")
                pcolor.Cyan(f"    - Windows: $env:{varName}='{varName}_value'")
                secret_path = ''
            if '~' in secret_path: secret_path = os.path.expanduser(secret_path)
            if not secret_path == '':
                if not os.path.isfile(secret_path): prRed(f'\n{"-"*91}\n\n  !!!Error!!! intersight_secret_key not found.')
                else:
                    secret_file = open(secret_path, 'r'); count = 0
                    if '-----BEGIN RSA PRIVATE KEY-----' in secret_file.read(): count += 1
                    secret_file.seek(0)
                    if '-----END RSA PRIVATE KEY-----' in secret_file.read(): count += 1
                    if count == 2: kwargs.args.intersight_secret_key = secret_path; secret_loop = True; valid = True
                    else: prRed(f'\n{"-"*91}\n\n  !!!Error!!! intersight_secret_key does not seem to contain a Valid Secret Key.')
            if not valid == True:
                kwargs.jdata = DotMap(
                    type = "string", minLength = 2, maxLength = 1024, pattern = '.*', title = 'Intersight',
                    description= 'Intersight Secret Key File Location.',
                    default    = f'{kwargs.home}{os.sep}Downloads{os.sep}SecretKey.txt')
                secret_path = variablePrompt(kwargs)

    #==============================================
    # Prompt User for Intersight FQDN
    #==============================================
    valid = False
    while valid == False:
        varValue = kwargs.args.intersight_fqdn
        if not varValue == None:
            varName = 'Intersight FQDN'
            if re.search(r'^[a-zA-Z0-9]{1,4}:', varValue): valid = validating.ip_address(varName, varValue)
            elif re.search(r'[a-zA-Z]', varValue): valid = validating.dns_name(varName, varValue)
            elif re.search(r'^([0-9]{1,3}\.){3}[0-9]{1,3}$', varValue): valid = validating.ip_address(varName, varValue)
            else: prRed(f'\n{"-"*91}\n\n  "{varValue}" is not a valid address.\n\n{"-"*91}\n')
        if valid == False:
            kwargs.jdata = kwargs.ezdata.ntp.allOf[1].properties.ntp_servers['items']
            kwargs.jdata.update(DotMap(description = 'Hostname of the Intersight FQDN',
                                       default = 'intersight.com', title = 'Intersight FQDN'))
            kwargs.args.intersight_fqdn = variablePrompt(kwargs)
            valid = True
    # Return kwargs
    return kwargs

#======================================================
# Function - Print with json dumps
#======================================================
def jprint(jDict): pcolor.LightGray(json.dumps(jDict, indent=4))

#======================================================
# Function - Load Previous YAML Files
#======================================================
def load_previous_configurations(kwargs):
    ezvars   = kwargs.ezdata.ezimm_class.properties
    vclasses = kwargs.ezdata.ezimm_class.properties.classes.enum
    dir_check= 0
    if os.path.isdir(kwargs.args.dir):
        dir_list = os.listdir(kwargs.args.dir)
        for i in dir_list:
            if i == 'templates':  dir_check += 1
            elif i == 'policies': dir_check += 1
            elif i == 'pools':    dir_check += 1
            elif i == 'profiles': dir_check += 1
            elif i == 'wizard': dir_check += 1
    if dir_check > 0:
        for item in vclasses:
            dest_dir = ezvars[item].directory
            if os.path.isdir(os.path.join(kwargs.args.dir, dest_dir)):
                dir_list = os.listdir(os.path.join(kwargs.args.dir, dest_dir))
                for i in dir_list:
                    if os.path.isfile(os.path.join(kwargs.args.dir, dest_dir, i)):
                        if re.search('.*yaml$', i):
                            yfile = open(os.path.join(kwargs.args.dir, dest_dir, i), 'r')
                            data = yaml.safe_load(yfile)
                            if not data == None:
                                for key, value in data.items():
                                    if not kwargs.imm_dict.orgs.get(key): kwargs.imm_dict.orgs[key] = {}
                                    for k, v in value.items():
                                        if not kwargs.imm_dict.orgs[key].get(k): kwargs.imm_dict.orgs[key][k] = {}
                                        kwargs.imm_dict.orgs[key][k].update(deepcopy(v))
    # Return kwargs
    return kwargs

#======================================================
# Function - Local User Policy - Users
#======================================================
def local_users_function(kwargs):
    loop_count = 1
    kwargs.local_users = []
    valid_config = False
    while valid_config == False:
        #==============================================
        # Loop Through Local User Atttributes
        #==============================================
        attributes    = DotMap()
        attribute_list= list(kwargs.ezdata['local_user.users'].properties.keys())
        attribute_list.remove('password')
        for e in attribute_list:
            kwargs.jdata = kwargs.ezdata['local_user.users'].properties[e]
            kwargs.jdata.multi_select = False
            attributes[e] = variablePrompt(kwargs)
        if kwargs.enforce_strong_password == True:
            print_with_textwrap(kwargs.ezdata['local_user.password_properties'].enforce_strong_password.description)
        kwargs.sensitive_var = f'local_user_password_{loop_count}'
        kwargs = sensitive_var_value(kwargs)
        attributes.password = loop_count
        attributes = DotMap(dict(sorted(attributes.toDict().items())))
        #==============================================
        # Show User Configuration
        #==============================================
        pcolor.Green(f'\n{"-"*91}\n')
        pcolor.Green(textwrap.indent(yaml.dump(attributes, Dumper=MyDumper, default_flow_style=False), " "*3, predicate=None))
        pcolor.Green(f'\n{"-"*91}\n')
        #======================================================================
        # * Prompt User to Accept Configuration, If Accepted add to Dictionary
        # * If User Selects to, Configure Additional
        #======================================================================
        kwargs.yes_or_no  = 'N'
        kwargs.policy_type= 'Local User'
        kwargs = exit_confirm_loop(kwargs)
        if kwargs.accept_configuration == True: kwargs.local_users.append(attributes)
        loop_count += 1
        valid_config = kwargs.configure_additional
    return kwargs

#======================================================
# Function - Merge Easy IMM Repository to Dest Folder
#======================================================
def merge_easy_imm_repository(kwargs):
    # Download the Easy IMM Comprehensive Example Base Repo
    baseRepo= kwargs.args.dir
    tfe_dir = 'tfe_modules'
    git_url = "https://github.com/terraform-cisco-modules/easy-imm-comprehensive-example"
    if not os.path.isdir(tfe_dir): os.mkdir(tfe_dir); Repo.clone_from(git_url, tfe_dir)
    if not os.path.isfile(os.path.join(tfe_dir, 'README.md')): Repo.clone_from(git_url, tfe_dir)
    else: g = cmd.Git(tfe_dir); g.pull()
    if not os.path.isdir(baseRepo): os.mkdir(baseRepo)
    # Now Loop over the folders and merge the module files
    for folder in ['defaults', '']:
        if folder == 'defaults':
            dest_dir = os.path.join(baseRepo, folder); src_dir = os.path.join(tfe_dir, 'defaults')
            if not os.path.isdir(dest_dir): os.mkdir(dest_dir)
        else: dest_dir = os.path.join(baseRepo); src_dir = os.path.join(tfe_dir)
        copy_files = os.listdir(src_dir)
        for fname in copy_files:
            if not os.path.isdir(os.path.join(src_dir, fname)): shutil.copy2(os.path.join(src_dir, fname), dest_dir)

#======================================================
# Function - Message for Invalid List Selection
#======================================================
def message_invalid_selection():
    prRed(f'\n{"-"*91}\n\n  !!!Error!!! Invalid Selection.  Please Select a valid Option from the List.\n\n{"-"*91}\n')

#======================================================
# Function - Message for Invalid Selection Y or N
#======================================================
def message_invalid_y_or_n(length):
    if length == 'short': dashRep = '-'*54
    else: dashRep = '-'*91
    prRed(f'\n{dashRep}\n\n  !!!Error!!! Invalid Value.  Please enter `Y` or `N`.\n\n{dashRep}\n')

#======================================================
# Function - Message Invalid FCoE VLAN
#======================================================
def message_fcoe_vlan(fcoe_id, vlan_policy):
    prRed(f'\n{"-"*91}\n\n  !!!Error!!!\n  The FCoE VLAN `{fcoe_id}` is already assigned to the VLAN Policy')
    prRed(f'  {vlan_policy}.  Please choose a VLAN id that is not already in use.\n\n{"-"*91}\n')

#======================================================
# Function(s) - Message Invalid Native VLAN
#======================================================
def message_invalid_native_vlan(nativeVlan, VlanList):
    prRed(f'\n{"-"*91}\n\n  !!!Error!!!\n  The Native VLAN `{nativeVlan}` was not in the VLAN Policy List.')
    prRed(f'  VLAN Policy List is: "{VlanList}"\n\n{"-"*91}\n')

#======================================================
# Function - Message Invalid VLAN/VSAN
#======================================================
def message_invalid_vxan():
    prRed(f'\n{"-"*91}\n\n  !!!Error!!!\n  Invalid Entry.  Please Enter a valid ID in the range of 1-4094.\n\n{"-"*91}\n')

#======================================================
# Function - Message Invalid VLAN
#======================================================
def message_invalid_vsan_id(vsan_policy, vsan_id, vsan_list):
    prRed(f'\n{"-"*91}\n\n  !!!Error!!!\n  The VSAN `{vsan_id}` is not in the VSAN Policy `{vsan_policy}`.')
    prRed(f'  Options are: {vsan_list}.\n\n{"-"*91}\n')

#======================================================
# Function - Message Starting Over
#======================================================
def message_starting_over(policy_type):  pcolor.Cyan(f'\n{"-"*54}\n\n  Starting `{policy_type}` Section over.\n\n{"-"*54}\n')

#======================================================
# Function - Change Policy Description to Sentence
#======================================================
def mod_pol_description(pol_description):
    pdescr = str.title(pol_description.replace('_', ' '))
    pdescr = (((pdescr.replace('Ipmi', 'IPMI')).replace('Ip', 'IP')).replace('Iqn', 'IQN')).replace('Ldap', 'LDAP')
    pdescr = (((pdescr.replace('Ntp', 'NTP')).replace('Sd', 'SD')).replace('Smtp', 'SMTP')).replace('Snmp', 'SNMP')
    pdescr = (((pdescr.replace('Ssh', 'SSH')).replace('Wwnn', 'WWNN')).replace('Wwpn', 'WWPN')).replace('Vsan', 'VSAN')
    return pdescr.replace('Vlan', 'VLAN')

#======================================================
# Function - Change Policy Description to Sentence
#======================================================
def name_prefix_suffix(policy, kwargs):
    name_prefix = ''
    name_suffix = ''
    if re.search('^ip|iqn|mac|resource|uuid|wwnn|wwpn$', policy): ptype = 'pools'
    elif re.search('chassis|domain|server|server_template', policy): ptype = 'profiles'
    else: ptype = 'policies'
    if kwargs.imm_dict.orgs[kwargs.org][ptype].get('name_prefix'):
        if kwargs.imm_dict.orgs[kwargs.org][ptype].name_prefix.get(policy):
            if len(kwargs.imm_dict.orgs[kwargs.org][ptype].name_prefix[policy]) > 0:
                name_prefix = kwargs.imm_dict.orgs[kwargs.org][ptype].name_prefix[policy]
    if name_prefix == '':
        if kwargs.imm_dict.orgs[kwargs.org][ptype].get('name_prefix'):
            if kwargs.imm_dict.orgs[kwargs.org][ptype].name_prefix.get('default'):
                if len(kwargs.imm_dict.orgs[kwargs.org][ptype].name_prefix['default']) > 0:
                    name_prefix = kwargs.imm_dict.orgs[kwargs.org][ptype].name_prefix['default']
    if kwargs.imm_dict.orgs[kwargs.org][ptype].get('name_suffix'):
        if kwargs.imm_dict.orgs[kwargs.org][ptype].name_suffix.get(policy):
            if len(kwargs.imm_dict.orgs[kwargs.org][ptype].name_suffix[policy]) > 0:
                name_suffix = kwargs.imm_dict.orgs[kwargs.org][ptype].name_suffix[policy]
    if name_suffix == '':
        if kwargs.imm_dict.orgs[kwargs.org][ptype].get('name_suffix'):
            if kwargs.imm_dict.orgs[kwargs.org][ptype].name_suffix.get('default'):
                if len(kwargs.imm_dict.orgs[kwargs.org][ptype].name_suffix['default']) > 0:
                    name_suffix = kwargs.imm_dict.orgs[kwargs.org][ptype].name_suffix['default']
    return name_prefix, name_suffix

#======================================================
# Function - Naming Rule
#======================================================
def naming_rule(name_prefix, name_suffix, org):
    if not name_prefix == '':  name = f'{name_prefix}_{name_suffix}'
    else: name = f'{org}_{name_suffix}'
    return name

#======================================================
# Function - Naming Rule Fabric Policy
#======================================================
def naming_rule_fabric(loop_count, name_prefix, org):
    letter = chr(ord('@')+loop_count+1)
    if not name_prefix == '':   name = f'{name_prefix}-{letter.lower()}'
    elif not org == 'default':  name = f'{org}-{letter.lower()}'
    else: name = f'fabric-{letter.lower()}'
    return name

#======================================================
# Function - Get Policies from Dictionary
#======================================================
def policies_parse(ptype, policy_type, kwargs):
    org  = kwargs.org
    kwargs.policies = []
    if not kwargs.imm_dict.orgs[org].get(ptype) == None:
        if not kwargs.imm_dict.orgs[org][ptype].get(policy_type) == None:
            kwargs.policies = {policy_type:kwargs.imm_dict.orgs[org][ptype][policy_type]}
        else: kwargs.policies = {policy_type:{}}
    else: kwargs.policies = {policy_type:{}}
    return kwargs

#======================================================
# Function - Print with Textwrap
#======================================================
def print_with_textwrap(description):
    pcolor.LightPurple(f'\n{"-"*108}\n')
    if '\n' in description:
        new_descr = description.split('\n')
        for line in new_descr:
            if '* ' in line: pcolor.LightGray(textwrap.fill(f'{line}',width=104, subsequent_indent='    '))
            elif '  - ' in line: pcolor.LightGray(textwrap.fill(f'{line}',width=104, subsequent_indent='      '))
            else: pcolor.LightGray(textwrap.fill(f'{line}',104, subsequent_indent=' '))
    else: pcolor.LightGray(textwrap.fill(f'{description}',104, subsequent_indent=' '))
    pcolor.LightPurple(f'\n{"-"*108}\n')

#======================================================
# Function - Validate input for each method
#======================================================
def process_kwargs(kwargs):
    # Validate User Input
    json_data = kwargs['validateData']
    validate_args(kwargs)
    error_count = 0
    error_list = []
    optional_args = json_data['optional_args']
    required_args = json_data['required_args']
    for item in required_args:
        if item not in kwargs['var_dict'].keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n'\
            ' - The Following REQUIRED Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    error_count = 0
    error_list = []
    for item in optional_args:
        if item not in kwargs['var_dict'].keys():
            error_count =+ 1
            error_list += [item]
    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n'\
            ' - The Following Optional Key(s) Were Not Found in kwargs: "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    # Load all required args values from kwargs
    error_count = 0
    error_list = []
    for item in kwargs['var_dict']:
        if item in required_args.keys():
            required_args[item] = kwargs['var_dict'][item]
            if required_args[item] == None:
                error_count =+ 1
                error_list += [item]

    if error_count > 0:
        error_ = '\n\n***Begin ERROR***\n\n'\
            ' - The Following REQUIRED Key(s) Argument(s) are Blank:\nPlease Validate "%s"\n\n****End ERROR****\n' % (error_list)
        raise InsufficientArgs(error_)

    for item in kwargs['var_dict']:
        if item in optional_args.keys():
            optional_args[item] = kwargs['var_dict'][item]
    # Combine option and required dicts for Jinja template render
    polVars = {**required_args, **optional_args}
    return(polVars)

#======================================================
# Function - Read Excel Workbook Data
#======================================================
def read_in(excel_workbook, kwargs):
    try:
        kwargs['wb'] = load_workbook(excel_workbook)
        pcolor.Cyan("Workbook Loaded.")
    except Exception as e:
        prRed(f'\n{"-"*91}\n\n  Something went wrong while opening the workbook - {excel_workbook}... ABORT!\n\n{"-"*91}\n')
        sys.exit(e)
    return kwargs

#======================================================
# Validate File is in Repo URL
#======================================================
def repo_url_test(file, pargs):
    repo_url = f'https://{pargs.repository_server}{pargs.repository_path}{file}'
    try:
        r = requests.head(repo_url, allow_redirects=True, verify=False, timeout=10)
    except requests.RequestException as e:
        prRed(f'\n{"-"*91}\n\n!!! ERROR !!!\n  Exception when calling {repo_url}:\n {e}\n\n{"-"*91}\n')
        sys.exit(1)
    return repo_url

#======================================================
# Function - Prompt User for Sensitive Values
#======================================================
def sensitive_var_value(kwargs):
    sensitive_var = kwargs.sensitive_var
    #=======================================================================================================
    # Check to see if the Variable is already set in the Environment, and if not prompt the user for Input.
    #=======================================================================================================
    if os.environ.get(sensitive_var) is None:
        pcolor.Cyan(f'\n{"-"*91}\n')
        pcolor.Cyan(f'  The Script did not find {sensitive_var} as an `environment` variable.')
        pcolor.Cyan(f'  To not be prompted for the value of `{kwargs.sensitive_var}` each time')
        pcolor.Cyan(f'  add the following to your local environemnt:\n')
        pcolor.Cyan(f"    - Linux: export {sensitive_var}='{kwargs.sensitive_var}_value'")
        pcolor.Cyan(f"    - Windows: $env:{sensitive_var}='{kwargs.sensitive_var}_value'")
        pcolor.Cyan(f'\n{"-"*91}\n')
    if os.environ.get(sensitive_var) is None and kwargs.sensitive_var == 'ipmi_key':
        pcolor.Cyan(f'\n{"-"*91}\n\n  The ipmi_key Must be in Hexidecimal Format [a-fA-F0-9]')
        pcolor.Cyan(f'  and no longer than 40 characters.\n\n{"-"*91}\n')
    if os.environ.get(sensitive_var) is None:
        valid = False
        while valid == False:
            varValue = input('press enter to continue: ')
            if varValue == '': valid = True
        valid = False
        while valid == False:
            if kwargs.get('multi_line_input'):
                pcolor.LightGray(f"Enter the value for {kwargs.sensitive_var}:")
                lines = []
                while True:
                    line = stdiomask.getpass(prompt='')
                    if line: lines.append(line)
                    else: break
                if not re.search('(certificate|private_key)', sensitive_var): secure_value = '\\n'.join(lines)
                else: secure_value = '\n'.join(lines)
            else:
                valid_pass = False
                while valid_pass == False:
                    password1 = stdiomask.getpass(prompt=f"Enter the value for {kwargs.sensitive_var}: ")
                    password2 = stdiomask.getpass(prompt=f"Re-Enter the value for {kwargs.sensitive_var}: ")
                    if password1 == password2: secure_value = password1; valid_pass = True
                    else: prRed('!!! ERROR !!! Sensitive Values did not match.  Please re-enter...')

            # Validate Sensitive Passwords
            cert_regex = re.compile(r'^\-{5}BEGIN (CERTIFICATE|PRIVATE KEY)\-{5}.*\-{5}END (CERTIFICATE|PRIVATE KEY)\-{5}$')
            sattributes = kwargs.ezdata.sensitive_variables
            if re.search('(certificate|private_key)', sensitive_var):
                if not re.search(cert_regex, secure_value): valid = True
                else:
                    prRed(f'\n{"-"*91}\n')
                    prRed(f'    !!! ERROR !!!\n  Invalid Value for the {sensitive_var}.  Please re-enter the {sensitive_var}.')
                    prRed(f'\n{"-"*91}\n')
            elif re.search('intersight_api_key_id', sensitive_var):
                kwargs.jdata = kwargs.ezdata.sensitive_variables.intersight_api_key_id
                valid = validate_sensitive(secure_value, kwargs)
            elif 'bind' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.ldap_binding_password
                valid = validate_sensitive(secure_value, kwargs)
            elif 'community' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.snmp_community_string
                valid = validate_sensitive(secure_value, kwargs)
            elif 'ipmi_key' in sensitive_var: valid = validating.ipmi_key_check(secure_value)
            elif 'iscsi_boot' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.iscsi_boot_password
                valid = validate_sensitive(secure_value, kwargs)
            elif 'local_user_password' in sensitive_var or 'ucs_password' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.local_user_password
                if kwargs.enforce_strong_password == True:
                    kwargs.jdata.maxLength = 20
                    valid = validate_strong_password(secure_value, kwargs)
                else: valid = validate_sensitive(secure_value, kwargs)
            elif 'persistent_passphrase' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.persistent_passphrase
                valid = validate_sensitive(secure_value, kwargs)
            elif 'snmp' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.snmp_password
                valid = validate_sensitive(secure_value, kwargs)
            elif 'vmedia' in sensitive_var:
                kwargs.jdata = kwargs.ezdata.sensitive_variables.vmedia_password
                valid = validate_sensitive(secure_value, kwargs)

        # Add Policy Variables to imm_dict
        if kwargs.get('org'):
            org = kwargs.org
            if not kwargs.imm_dict.orgs.get(org):
                kwargs.imm_dict.orgs[org] = DotMap()
                if not kwargs.imm_dict.orgs[org].get('sensitive_vars'):
                    kwargs.imm_dict.orgs[org].sensitive_vars = []
                kwargs.imm_dict.orgs[org].sensitive_vars.append(sensitive_var)

        # Add the Variable to the Environment
        os.environ[sensitive_var] = '%s' % (secure_value)
        kwargs.var_value = secure_value
    else:
        # Add the Variable to the Environment
        if not kwargs.get('multi_line_input'): kwargs.var_value = os.environ.get(sensitive_var)
        else: kwargs.var_value = (os.environ.get(sensitive_var)).replace('\n', '\\n')
    return kwargs

#======================================================
# Function - Wizard for SNMP Trap Servers
#======================================================
def snmp_trap_servers(kwargs):
    loop_count = 1
    kwargs.snmp_traps = []
    valid_config = False
    while valid_config == False:
        #==============================================
        # Loop Through SNMP Trap Server Atttributes
        #==============================================
        attributes= DotMap()
        attribute_list = list(kwargs.ezdata['snmp.snmp_trap_destinations'].properties.keys())
        if len(kwargs.snmp_users) > 0:
            for e in ['community_string', 'trap_type', 'user']: attribute_list.remove(e)
        else: attribute_list.pop('user')
        for e in attribute_list:
            kwargs.jdata = kwargs.ezdata['snmp.snmp_trap_destinations'].properties[e]
            kwargs.jdata.multi_select = False
            attributes[e] = variablePrompt(kwargs)
        if len(kwargs.snmp_users) > 0:
            kwargs.jdata = kwargs.ezdata['snmp.snmp_trap_destinations'].properties['user']
            kwargs.jdata.enum = [e.name for e in kwargs.snmp_users]
            kwargs.jdata.multi_select = False
            attributes.user = variablePrompt(kwargs)
        else:
            kwargs.jdata = kwargs.ezdata['snmp.snmp_trap_destinations'].properties['trap_type']
            kwargs.jdata.multi_select = False
            attributes.trap_type = variablePrompt(kwargs)
            kwargs.sensitive_var = f'snmp_trap_community_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
            attributes.community_string = loop_count
        attributes = DotMap(dict(sorted(attributes.toDict().items())))
        #==============================================
        # Show User Configuration
        #==============================================
        pcolor.Green(f'\n{"-"*91}\n')
        pcolor.Green(textwrap.indent(yaml.dump(attributes, Dumper=MyDumper, default_flow_style=False), " "*3, predicate=None))
        pcolor.Green(f'\n{"-"*91}\n')
        #======================================================================
        # * Prompt User to Accept Configuration, If Accepted add to Dictionary
        # * If User Selects to, Configure Additional
        #======================================================================
        kwargs.yes_or_no  = 'N'
        kwargs.policy_type= 'SNMP Trap Destination'
        kwargs = exit_confirm_loop(kwargs)
        if kwargs.accept_configuration == True: kwargs.snmp_traps.append(attributes)
        loop_count += 1
        valid_config = kwargs.configure_additional
    return kwargs

#======================================================
# Function - Wizard for SNMP Users
#======================================================
def snmp_users(kwargs):
    loop_count = 1
    kwargs.snmp_users = []
    valid_users = False
    while valid_users == False:
        #==============================================
        # Loop Through SNMP User Atttributes
        #==============================================
        attributes    = DotMap()
        attribute_list= list(kwargs.ezdata['snmp.snmp_users'].properties.keys())
        for e in ['auth_password', 'privacy_password', 'privacy_type']: attribute_list.remove(e)
        for e in attribute_list:
            kwargs.jdata = kwargs.ezdata['snmp.snmp_users'].properties[e]
            attributes[e] = variablePrompt(kwargs)
        if attributes.security_level == 'AuthPriv':
            kwargs.jdata = kwargs.ezdata['snmp.snmp_users'].properties.privacy_type
            attributes.privacy_type = variablePrompt(kwargs)
        if re.search('Auth(No)?Priv', attributes.security_level):
            kwargs.sensitive_var = f'snmp_auth_password_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
            attributes.auth_password = loop_count
        if re.search('AuthPriv', attributes.security_level):
            kwargs.sensitive_var = f'snmp_privacy_password_{loop_count}'
            kwargs = sensitive_var_value(kwargs)
            attributes.privacy_password = loop_count
        attributes = DotMap(dict(sorted(attributes.toDict().items())))
        #==============================================
        # Show User Configuration
        #==============================================
        pcolor.Green(f'\n{"-"*91}\n')
        pcolor.Green(textwrap.indent(yaml.dump(attributes, Dumper=MyDumper, default_flow_style=False), " "*3, predicate=None))
        pcolor.Green(f'\n{"-"*91}\n')
        #======================================================================
        # * Prompt User to Accept Configuration, If Accepted add to Dictionary
        # * If User Selects to, Configure Additional
        #======================================================================
        kwargs.yes_or_no  = 'N'
        kwargs.policy_type= 'SNMP User'
        kwargs = exit_confirm_loop(kwargs)
        if kwargs.accept_configuration == True: kwargs.snmp_users.append(attributes)
        loop_count += 1
        valid_users = kwargs.configure_additional
    return kwargs

#======================================================
# Function - Define stdout_log output
#======================================================
def stdout_log(ws, row_num):
    if log_level == 0: return
    elif ((log_level == (1) or log_level == (2)) and (ws) and (row_num is None)) and row_num == 'begin':
        pcolor.Cyan(f'\n{"-"*91}\n\n   Begin Worksheet "{ws.title}" evaluation...\n\n{"-"*91}\n')
    elif (log_level == (1) or log_level == (2)) and row_num == 'end':
        pcolor.Cyan(f'\n{"-"*91}\n\n   Completed Worksheet "{ws.title}" evaluation...\n\n{"-"*91}\n')
    elif log_level == (2) and (ws) and (row_num is not None):
        pcolor.Cyan(f'    - Evaluating Row{" "*(4-len(row_num))}{row_num}...')
    else: return

#======================================================
# Function - Wizard for Syslog Servers
#======================================================
def syslog_servers(kwargs):
    kwargs.remote_logging = []
    loop_count = 0
    valid_config = False
    while valid_config == False:
        if loop_count < 2:
            #==============================================
            # Loop Through SNMP Trap Server Atttributes
            #==============================================
            attributes= DotMap()
            attribute_list = list(kwargs.ezdata['syslog.remote_logging'].properties.keys())
            for e in attribute_list:
                kwargs.jdata = kwargs.ezdata['syslog.remote_logging'].properties[e]
                kwargs.jdata.multi_select = False
                attributes[e] = variablePrompt(kwargs)
            #==============================================
            # Show User Configuration
            #==============================================
            pcolor.Green(f'\n{"-"*91}\n')
            pcolor.Green(textwrap.indent(yaml.dump(attributes, Dumper=MyDumper, default_flow_style=False), " "*3, predicate=None))
            pcolor.Green(f'\n{"-"*91}\n')
            #======================================================================
            # * Prompt User to Accept Configuration, If Accepted add to Dictionary
            # * If User Selects to, Configure Additional
            #======================================================================
            if loop_count == 0: kwargs.yes_or_no  = 'Y'
            else: kwargs.yes_or_no  = 'N'
            kwargs.policy_type= 'Syslog Remote Servers'
            kwargs = exit_confirm_loop(kwargs)
            if kwargs.accept_configuration == True: kwargs.remote_logging.append(attributes)
            loop_count += 1
            valid_config = kwargs.configure_additional
        else: valid_config = True
    return kwargs

#======================================================
# Function - Create a List of Subnet Hosts
#======================================================
def subnet_list(kwargs):
    if kwargs.ip_version == 'v4': prefix = kwargs.subnetMask
    else: prefix = kwargs.prefix
    gateway = kwargs.defaultGateway
    return list(ipaddress.ip_network(f"{gateway}/{prefix}", strict=False).hosts())

#======================================================
# Function - Format Terraform Files
#======================================================
def terraform_fmt(folder):
    # Run terraform fmt to cleanup the formating for all of the auto.tfvar files and tf files if needed
    pcolor.Cyan(f'\n-----------------------------------------------------------------------------\n')
    pcolor.Cyan(f'  Running "terraform fmt" in folder "{folder}",')
    pcolor.Cyan(f'  to correct variable formatting!')
    pcolor.Cyan(f'\n-----------------------------------------------------------------------------\n')
    p = subprocess.Popen(['terraform', 'fmt', folder], stdout = subprocess.PIPE, stderr = subprocess.PIPE)
    pcolor.Cyan('Format updated for the following Files:')
    for line in iter(p.stdout.readline, b''):
        line = line.decode("utf-8")
        line = line.strip()
        pcolor.Cyan(f'- {line}')

#========================================================
# Function to Pull Latest Versions of Providers
#========================================================
def terraform_provider_config(kwargs):
    url_list = [
        'https://github.com/CiscoDevNet/terraform-provider-intersight/tags/',
        'https://github.com/hashicorp/terraform/tags',
        'https://github.com/netascode/terraform-provider-utils/tags/']
    for url in url_list:
        # Get the Latest Release Tag for the Provider
        r = requests.get(url, stream=True)
        repoVer = 'BLANK'
        stringMatch = False
        while stringMatch == False:
            for line in r.iter_lines():
                toString = line.decode("utf-8")
                if re.search(r'/releases/tag/v(\d+\.\d+\.\d+)\"', toString):
                    repoVer = re.search('/releases/tag/v(\d+\.\d+\.\d+)', toString).group(1)
                    break
            stringMatch = True
        # Make sure the latest_versions Key exists
        if kwargs.get('latest_versions') == None: kwargs.latest_versions = {}
        # Set Provider Version
        if   'intersight' in url: kwargs.latest_versions.intersight_provider_version = repoVer
        elif 'netascode' in url:  kwargs.latest_versions.utils_provider_version = repoVer
        else: kwargs.latest_versions.terraform_version = repoVer
    # Return kwargs
    return kwargs
    
#======================================================
# Function - Prompt User for Sensitive Variables
#======================================================
def tfc_sensitive_variables(var_value, kwargs):
    polVars = DotMap(Variable = var_value)
    polVars.Description = ("".join(var_value.split('_'))).title()
    polVars.Description = mod_pol_description(polVars.Description)
    kwargs = sensitive_var_value(kwargs)
    polVars.varValue = kwargs.var_value
    polVars.varId = var_value
    polVars.varKey = var_value
    polVars.Sensitive = True
    pcolor.Cyan('* Adding "{}" to "{}"').format(polVars.description, kwargs.workspaceName)
    return polVars

#==========================================================
# Function - Prompt User for Chassis/Server Serial Numbers
#==========================================================
def ucs_serial(kwargs):
    baseRepo    = kwargs.args.dir
    device_type = kwargs.device_type
    org         = kwargs.org
    yaml_file   = kwargs.yaml_file
    valid = False
    while valid == False:
        pcolor.Cyan(f'\n-------------------------------------------------------------------------------------------\n')
        pcolor.Cyan(f'  Note: If you do not have the Serial Number(s) at this time you can manually add it to:')
        pcolor.Cyan(f'    - {baseRepo}{os.sep}{org}{os.sep}profiles{os.sep}{yaml_file}.yaml')
        pcolor.Cyan(f'      file later.')
        pcolor.Cyan(f'\n-------------------------------------------------------------------------------------------\n')
        serial = input(f'What is the Serial Number of the {device_type}? [press enter to skip]: ')
        if serial == '': serial = 'unknown'; valid = True
        elif re.fullmatch(r'^[A-Z]{3}[2-3][\d]([0][1-9]|[1-4][0-9]|[5][1-3])[\dA-Z]{4}$', serial): valid = True
        else:
            prRed(f'\n-------------------------------------------------------------------------------------------\n')
            prRed(f'  Error!! Invalid Serial Number.  "{serial}" is not a valid serial.')
            prRed(f'\n-------------------------------------------------------------------------------------------\n')
    return serial

#======================================================
# Function - Prompt User for Domain Serial Numbers
#======================================================
def ucs_domain_serials(kwargs):
    baseRepo = kwargs['args'].dir
    org = kwargs['org']
    pcolor.Cyan(f'\n-------------------------------------------------------------------------------------------\n')
    pcolor.Cyan(f'  Note: If you do not have the Serial Numbers at this time you can manually add them here:\n')
    pcolor.Cyan(f'    * {baseRepo}{os.sep}{org}{os.sep}profiles{os.sep}domain.yaml\n')
    pcolor.Cyan(f'  After the Wizard has completed.')
    pcolor.Cyan(f'\n-------------------------------------------------------------------------------------------\n')
    valid = False
    while valid == False:
        polVars = {}
        fabrics = ['A','B']
        for x in fabrics:
            polVars[f'serial_{x}'] = input(f'What is the Serial Number of Fabric {x}? [press enter to skip]: ')
            if polVars[f'serial_{x}'] == '':
                polVars[f'serial_{x}'] = 'unknown'
                valid = True
            elif re.fullmatch(r'^[A-Z]{3}[2-3][\d]([0][1-9]|[1-4][0-9]|[5][1-3])[\dA-Z]{4}$', polVars[f'serial_{x}']):
                valid = True
            else:
                prRed(f'\n-------------------------------------------------------------------------------------------\n')
                prRed('  Error!! Invalid Serial Number.  "{}" is not a valid serial.').format(polVars[f'serial_{x}'])
                prRed(f'\n-------------------------------------------------------------------------------------------\n')
    serials = [polVars['serial_A'], polVars['serial_B']]
    return serials

#========================================================
# Function to Validate Worksheet User Input
#========================================================
def validate_args(json_data, kwargs):
    json_data = kwargs['validateData']
    for i in json_data['required_args']:
        if json_data[i]['type'] == 'boolean':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                validating.boolean(i, kwargs)
        elif json_data[i]['type'] == 'hostname':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                if ':' in kwargs['var_dict'][i]: validating.ip_address_ws(i, kwargs)
                elif re.search('[a-z]', kwargs['var_dict'][i], re.IGNORECASE): validating.dns_name_ws(i, kwargs)
                else: validating.ip_address_ws(i, kwargs)
        elif json_data[i]['type'] == 'list_of_email':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for email in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = email
                    validating.email_ws(f'{i}_{count}', kwargs)
        elif json_data[i]['type'] == 'email':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''): validating.email_ws(i, kwargs)
        elif json_data[i]['type'] == 'integer':
            if kwargs['var_dict'][i] == None: kwargs['var_dict'][i] = json_data[i]['default']
            else: validating.number_check(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_domains':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for domain in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'domain_{count}'] = domain
                    validating.domain_ws(f'domain_{count}', kwargs)
                    kwargs['var_dict'].pop(f'domain_{count}')
                    count += 1
        elif json_data[i]['type'] == 'list_of_hosts':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                count = 1
                for hostname in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = hostname
                    if ':' in hostname: validating.ip_address_ws(f'{i}_{count}', kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE): validating.dns_name_ws(f'{i}_{count}', kwargs)
                    else: validating.ip_address_ws(f'{i}_{count}', kwargs)
                    kwargs['var_dict'].pop(f'{i}_{count}')
                    count += 1
        elif json_data[i]['type'] == 'list_of_integer':
            if kwargs['var_dict'][i] == None: kwargs['var_dict'][i] = json_data[i]['default']
            else: validating.number_list(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_string':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''): validating.string_list(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_values':
            if kwargs['var_dict'][i] == None: kwargs['var_dict'][i] = json_data[i]['default']
            else: validating.list_values(i, json_data, kwargs)
        elif json_data[i]['type'] == 'list_of_vlans':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''): validating.vlans(i, kwargs)
        elif json_data[i]['type'] == 'string':
            if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
                validating.string_pattern(i, json_data, kwargs)
        else: prRed(f"error validating.  Type not found {json_data[i]['type']}. 2."); sys.exit(1)
    for i in json_data['optional_args']:
        if not (kwargs['var_dict'][i] == None or kwargs['var_dict'][i] == ''):
            if re.search(r'^module_[\d]+$', i): validating.list_values_key('modules', i, json_data, kwargs)
            elif json_data[i]['type'] == 'boolean': validating.boolean(i, json_data, kwargs)
            elif json_data[i]['type'] == 'domain': validating.domain_ws(i, kwargs)
            elif json_data[i]['type'] == 'list_of_email':
                count = 1
                for email in kwargs['var_dict'][i].split(','):
                    kwargs['var_dict'][f'{i}_{count}'] = email
                    validating.email_ws(f'{i}_{count}', json_data, kwargs)
            elif json_data[i]['type'] == 'email': validating.email_ws(i, json_data, kwargs)
            elif json_data[i]['type'] == 'hostname':
                if ':' in kwargs['var_dict'][i]: validating.ip_address_ws(i, kwargs)
                elif re.search('[a-z]', kwargs['var_dict'][i], re.IGNORECASE): validating.dns_name_ws(i, kwargs)
                else: validating.ip_address_ws(i, kwargs)
            elif json_data[i]['type'] == 'integer': validating.number_check(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_integer': validating.number_list(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_hosts':
                count = 1
                for hostname in kwargs['var_dict'][i].split(','):
                    kwargs[f'{i}_{count}'] = hostname
                    if ':' in hostname: validating.ip_address_ws(f'{i}_{count}', kwargs)
                    elif re.search('[a-z]', hostname, re.IGNORECASE): validating.dns_name_ws(f'{i}_{count}', kwargs)
                    else: validating.ip_address_ws(f'{i}_{count}', kwargs)
                    kwargs['var_dict'].pop(f'{i}_{count}')
                    count += 1
            elif json_data[i]['type'] == 'list_of_macs':
                count = 1
                for mac in kwargs['var_dict'][i].split(','):
                    kwargs[f'{i}_{count}'] = mac
                    validating.mac_address(f'{i}_{count}', kwargs)
                    kwargs.pop(f'{i}_{count}')
                    count += 1
            elif json_data[i]['type'] == 'list_of_string': validating.string_list(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_values': validating.list_values(i, json_data, kwargs)
            elif json_data[i]['type'] == 'list_of_vlans':  validating.vlans(i, kwargs)
            elif json_data[i]['type'] == 'mac_address':    validating.mac_address(i, kwargs)
            elif json_data[i]['type'] == 'string':         validating.string_pattern(i, json_data, kwargs)
            else: prRed(f"error validating.  Type not found {json_data[i]['type']}. 3."); sys.exit(1)
    return kwargs

#======================================================
# Function - Check VLAN exists in VLAN Policy
#======================================================
def validate_vlan_in_policy(vlan_policy_list, vlan_id):
    valid = False
    while valid == False:
        vlan_count = 0
        for vlan in vlan_policy_list:
            if int(vlan_id) == 1: vlan_count = 1; continue
            if int(vlan) == int(vlan_id): vlan_count = 1; continue
        if vlan_count == 1: valid = True; return valid
        else:
            prRed(f'\n-------------------------------------------------------------------------------------------\n')
            prRed(f'  VLAN {vlan_id} not found in the VLAN Policy List.  Please us a VLAN from the list below:')
            prRed(f'  {vlan_policy_list}')
            prRed(f'\n-------------------------------------------------------------------------------------------\n')
            return valid

def validate_ipmi_key(varValue):
    valid_count = 0
    varValue = varValue.capitalize()
    if ((varValue < '0' or varValue > '9') and (varValue < 'A' or varValue > 'F')): valid_count += 1
    if not validators.length(varValue, min=2, max=40): valid_count += 1
    if not len(varValue) % 2 == 0: valid_count += 1
    if not valid_count == 0:
        prRed(f'\n-----------------------------------------------------------------------------\n')
        prRed(f'   Error with ipmi_key!!  The encryption key should have an even number of ')
        prRed(f'   hexadecimal characters and not exceed 40 characters.\n')
        prRed(f'   Valid Hex Characters are:')
        prRed(f'    - {string.hexdigits}')
        prRed(f'\n-----------------------------------------------------------------------------\n')
        return False
    else: return True

#======================================================
# Function - Validate Sensitive Strings
#======================================================
def validate_sensitive(secure_value, kwargs):
    invalid_count = 0
    if not validators.length(secure_value, min=int(kwargs.minLength), max=int(kwargs.maxLength)):
        invalid_count += 1
        prRed(f'\n--------------------------------------------------------------------------------------\n')
        prRed(f'   !!! {kwargs.sensitive_var} is Invalid!!!')
        prRed(f'   Length Must be between {kwargs.minLength} and {kwargs.maxLength} characters.')
        prRed(f'\n--------------------------------------------------------------------------------------\n')
    if not re.search(kwargs.pattern, secure_value):
        invalid_count += 1
        prRed(f'\n--------------------------------------------------------------------------------------\n')
        prRed(f'   !!! Invalid Characters in {kwargs.sensitive_var}.  The allowed characters are:')
        prRed(f'   - "{kwargs.pattern}"')
        prRed(f'\n--------------------------------------------------------------------------------------\n')
    if invalid_count == 0: return True
    else: return False

#======================================================
# Function - Validate Sensitive Strings
#======================================================
def validate_strong_password(secure_value, kwargs):
    invalid_count = 0; valid_count = 0
    if re.search(kwargs.username, secure_value, re.IGNORECASE): invalid_count += 1
    if not validators.length(str(secure_value), min=int(kwargs.jdata.minLength), max=int(kwargs.jdata.maxLength)):
        invalid_count += 1
    else: valid_count +=1
    if re.search(r'[a-z]', secure_value): valid_count += 1
    if re.search(r'[A-Z]', secure_value): valid_count += 1
    if re.search(r'[0-9]', secure_value): valid_count += 1
    if re.search(r'[\!\@\#\$\%\^\&\*\-\_\+\=]', secure_value): valid_count += 1
    if not invalid_count == 0 and valid_count >= 4:
        prRed(f'\n---------------------------------------------------------------------------------------\n')
        prRed(f"   Error with {kwargs.sensitive_var}! The password failed one of the following complexity rules:")
        prRed(f'     - The password must have a minimum of 8 and a maximum of 20 characters.')
        prRed(f"     - The password must not contain the User's Name.")
        prRed(f'     - The password must contain characters from three of the following four categories.')
        prRed(f'       * English uppercase characters (A through Z).')
        prRed(f'       * English lowercase characters (a through z).')
        prRed(f'       * Base 10 digits (0 through 9).')
        prRed(f'       * Non-alphabetic characters (! , @, #, $, %, ^, &, *, -, _, +, =)')
        prRed(f'\n---------------------------------------------------------------------------------------\n')
        return False
    else: return True

#======================================================
# Function - Prompt for Answer to Question from List
#======================================================
def variableFromList(kwargs):
    #==============================================
    # Set Function Variables
    #==============================================
    default     = kwargs.jdata.default
    description = kwargs.jdata.description
    title       = kwargs.jdata.title
    if not kwargs.jdata.get('multi_select'): kwargs.jdata.multi_select = False
    #==============================================
    # Sort the Variables
    #==============================================
    if kwargs.jdata.get('sort') == False: vars = kwargs.jdata.enum
    else: vars = sorted(kwargs.jdata.enum, key=str.casefold)
    valid = False
    while valid == False:
        pcolor.LightPurple(f'\n{"-"*108}\n')
        if '\n' in description:
            description = description.split('\n')
            for line in description:
                if '*' in line: pcolor.LightGray(textwrap.fill(f' {line}',width=104, subsequent_indent='    '))
                else: pcolor.LightGray(textwrap.fill(f'{line}',104, subsequent_indent=' '))
        else: pcolor.LightGray(textwrap.fill(f'{description}',104, subsequent_indent=' '))
        if kwargs.jdata.get('multi_select') == True:
            pcolor.Yellow(
                '\n     Note: Answer can be:\n       * Single: 1 or 5\n       * Multiple: `1,2,3` or `1-3,5-6` in example')
        if kwargs.jdata.get('multi_select') == True: pcolor.Yellow(f'    Select Option(s) Below:')
        else: pcolor.Yellow(f'\n    Select an Option Below:')
        for index, value in enumerate(vars):
            index += 1
            if value == default: default_index = index
            if index < 10: pcolor.Cyan(f'     {index}. {value}')
            else: pcolor.Cyan(f'    {index}. {value}')
        pcolor.LightPurple(f'\n{"-"*108}\n')
        if kwargs.jdata.get('multi_select') == True:
            if not default == '':
                var_selection   = input(f'Please Enter the Option Number(s) to select for {title}.  [{default_index}]: ')
            else: var_selection = input(f'Please Enter the Option Number(s) to select for {title}: ')
        else:
            if not default == '':
                var_selection   = input(f'Please Enter the Option Number to select for {title}.  [{default_index}]: ')
            else: var_selection = input(f'Please Enter the Option Number to select for {title}: ')
        if not default == '' and var_selection == '':
            var_selection = default_index
        if kwargs.jdata.multi_select == False and re.search(r'^[0-9]+$', str(var_selection)):
            for index, value in enumerate(vars):
                index += 1
                if int(var_selection) == index: selection = value; valid = True
        elif kwargs.jdata.multi_select == True and re.search(r'(^[0-9]+$|^[0-9\-,]+[0-9]$)', str(var_selection)):
            var_list = vlan_list_full(var_selection)
            var_length = int(len(var_list))
            var_count = 0
            selection = []
            for index, value in enumerate(vars):
                index += 1
                for vars in var_list:
                    if int(vars) == index: var_count += 1; selection.append(value)
            if var_count == var_length: valid = True
            else: prRed(f'\n{"-"*91}\n\n  The list of Vars {var_list} did not match the available list.\n\n{"-"*91}\n')
        if valid == False: message_invalid_selection()
    return selection, valid

#======================================================
# Function - Prompt User for Answer to Question
#======================================================
def variablePrompt(kwargs):
    #==============================================
    # Improper Value Notifications
    #==============================================
    def invalid_boolean(title, answer):
        prRed(f'\n{"-"*91}\n   `{title}` value of `{answer}` is Invalid!!! Please enter `Y` or `N`.\n{"-"*91}\n')
    def invalid_integer(title, answer):
        prRed(f'\n{"-"*91}\n   `{title}` value of `{answer}` is Invalid!!!  Valid range is `{minimum}-{maximum}`.\n{"-"*91}\n')
    def invalid_string(title, answer):
        prRed(f'\n{"-"*91}\n   `{title}` value of `{answer}` is Invalid!!!\n{"-"*91}\n')
    #==============================================
    # Set Function Variables
    #==============================================
    default     = kwargs.jdata.default
    description = kwargs.jdata.description
    optional    = False
    title       = kwargs.jdata.title
    #==============================================
    # Print `description` if not enum
    #==============================================
    if not kwargs.jdata.get('enum'):  print_with_textwrap(description)
    #==============================================
    # Prompt User for Answer
    #==============================================
    valid = False
    while valid == False:
        if kwargs.jdata.get('enum'):  answer, valid = variableFromList(kwargs)
        elif kwargs.jdata.type == 'boolean':
            if default == True: default = 'Y'
            else: default = 'N'
            answer = input(f'Enter `Y` for `True` or `N` for `False` for {title}. [{default}]:')
            if answer == '':
                if default == 'Y': answer = True
                elif default == 'N': answer = False
                valid = True
            elif answer == 'N': answer = False; valid = True
            elif answer == 'Y': answer = True;  valid = True
            else: invalid_boolean(title, answer)
        elif kwargs.jdata.type == 'integer':
            maximum = kwargs.jdata.maximum
            minimum = kwargs.jdata.minimum
            if kwargs.jdata.get('optional') == True:
                optional = True
                answer = input(f'Enter the value for {title} [press enter to skip]: ')
            else: answer = input(f'Enter the Value for {title}. [{default}]: ')
            if optional == True and answer == '': valid = True
            elif answer == '': answer = default
            if optional == False:
                if re.fullmatch(r'^[0-9]+$', str(answer)):
                    if kwargs.jdata.title == 'snmp_port':
                        valid = validating.snmp_port(title, answer, minimum, maximum)
                    else: valid = validating.number_in_range(title, answer, minimum, maximum)
                else: invalid_integer(title, answer)
        elif kwargs.jdata.type == 'string':
            if kwargs.jdata.get('optional') == True:
                optional = True
                answer = input(f'Enter the value for {title} [press enter to skip]: ')
            elif not default == '': answer = input(f'Enter the value for {title} [{default}]: ')
            else: answer = input(f'Enter the value for {title}: ')
            if optional == True and answer == '': valid = True
            elif answer == '': answer = default; valid = True
            elif not answer == '':
                maxLength = kwargs.jdata.maxLength
                minLength = kwargs.jdata.minLength
                pattern   = kwargs.jdata.pattern
                valid = validating.length_and_regex(answer, minLength, maxLength, pattern, title)
        else: invalid_string(title, answer)
    if kwargs.jdata.get('optional'): kwargs.jdata.pop('optional')
    if kwargs.jdata.get('multi_select'): kwargs.jdata.pop('multi_select')
    return answer

#======================================================
# Function - Collapse VLAN List
#======================================================
def vlan_list_format(vlan_list_expanded):
    vlan_list  = sorted(vlan_list_expanded)
    vlanGroups = itertools.groupby(vlan_list, key=lambda item, c=itertools.count():item-next(c))
    tempvlans  = [list(g) for k, g in vlanGroups]
    vlanList   = [str(x[0]) if len(x) == 1 else "{}-{}".format(x[0],x[-1]) for x in tempvlans]
    vlan_list  = ",".join(vlanList)
    return vlan_list

#======================================================
# Function - Expand VLAN List
#======================================================
def vlan_list_full(vlan_list):
    full_vlan_list = []
    if re.search(r',', str(vlan_list)):
        vlist = vlan_list.split(',')
        for v in vlist:
            if re.fullmatch('^\\d{1,4}\\-\\d{1,4}$', v):
                a,b = v.split('-'); a = int(a); b = int(b); vrange = range(a,b+1)
                for vl in vrange: full_vlan_list.append(int(vl))
            elif re.fullmatch('^\\d{1,4}$', v): full_vlan_list.append(int(v))
    elif re.search('\\-', str(vlan_list)):
        a,b = vlan_list.split('-'); a = int(a); b = int(b); vrange = range(a,b+1)
        for v in vrange: full_vlan_list.append(int(v))
    else: full_vlan_list.append(vlan_list)
    return full_vlan_list

#======================================================
# Function - To Request Native VLAN
#======================================================
def vlan_native_function(vlan_policy_list, vlan_list):
    native_count = 0
    nativeVlan = ''
    nativeValid = False
    while nativeValid == False:
        nativeVlan = input('Do you want to Configure one of the VLANs as a Native VLAN?  [press enter to skip]:')
        if nativeVlan == '': nativeValid = True
        else:
            for vlan in vlan_policy_list:
                if int(nativeVlan) == int(vlan): native_count = 1; break
            if not native_count == 1: message_invalid_native_vlan(nativeVlan, vlan_list)
            else: nativeValid = True
    return nativeVlan

#======================================================
# Function - Prompt for VLANs and Configure Policy
#======================================================
def vlan_pool(name):
    valid = False
    while valid == False:
        pcolor.Cyan(f'\n{"-"*91}\n')
        pcolor.Cyan(f'  The allowed vlan list can be in the format of:')
        pcolor.Cyan(f'     5 - Single VLAN\n     1-10 - Range of VLANs\n     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
        pcolor.Cyan(f'     1-10,20-30 - Ranges and Lists of VLANs\n\n{"-"*91}\n')
        VlanList = input(f'Enter the VLAN or List of VLANs to assign to the Domain VLAN Pool {name}: ')
        if not VlanList == '':
            vlanListExpanded = vlan_list_full(VlanList)
            valid_vlan = True
            for vlan in vlanListExpanded:
                valid_vlan = validating.number_in_range('VLAN ID', vlan, 1, 4094)
                if valid_vlan == False: break
            if valid_vlan == False:
                prRed(f'\n{"-"*91}\n')
                prRed(f'  !!!Error!!!\n  With VLAN(s) assignment. VLAN List: "{VlanList}" is not Valid.')
                prRed(f'  The allowed vlan list can be in the format of:')
                prRed(f'     5 - Single VLAN\n     1-10 - Range of VLANs\n     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
                prRed(f'     1-10,20-30 - Ranges and Lists of VLANs\n\n{"-"*91}\n')
            else: valid = True
        else:
            prRed(f'\n{"-"*91}\n')
            prRed(f'  The allowed vlan list can be in the format of:')
            prRed(f'     5 - Single VLAN\n     1-10 - Range of VLANs\n     1,2,3,4,5,11,12,13,14,15 - List of VLANs')
            prRed(f'     1-10,20-30 - Ranges and Lists of VLANs\n\n{"-"*91}\n')
    return VlanList,vlanListExpanded

#========================================================
# Function to Determine which sites to write files to.
#========================================================
def write_to_repo_folder(polVars, kwargs):
    baseRepo   = kwargs.args.dir
    dest_file  = kwargs.dest_file

    # Setup jinja2 Environment
    template_path = pkg_resources.resource_filename(f'policies', 'templates/')
    templateLoader = jinja2.FileSystemLoader(searchpath=(template_path + 'provider/'))
    templateEnv = jinja2.Environment(loader=templateLoader)

    # Define the Template Source
    template = templateEnv.get_template(kwargs.template_file)

    # Make sure the Destination Path and Folder Exist
    if not os.path.isdir(os.path.join(baseRepo)): dest_path = f'{os.path.join(baseRepo)}'; os.makedirs(dest_path)
    dest_dir = os.path.join(baseRepo)
    if not os.path.exists(os.path.join(dest_dir, dest_file)):
        create_file = f'type nul >> {os.path.join(dest_dir, dest_file)}'; os.system(create_file)
    tf_file = os.path.join(dest_dir, dest_file)
    wr_file = open(tf_file, 'w')

    # Render Payload and Write to File
    polVars = json.loads(json.dumps(polVars))
    polVars = {'keys':polVars}
    payload = template.render(polVars)
    wr_file.write(payload)
    wr_file.close()
